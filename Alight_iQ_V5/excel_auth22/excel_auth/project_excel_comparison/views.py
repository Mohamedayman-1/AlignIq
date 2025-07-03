from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from .models import User, FileUpload, Comparison, Database_Connection, Database_Comparison
from .serializers import RegisterSerializer, LoginSerializer, FileUploadSerializer, ComparisonSerializer, DatabaseConnectionSerializer
from rest_framework_simplejwt.tokens import RefreshToken
from datetime import timedelta, datetime, timezone as dt_timezone
from django.utils import timezone
from .permissions import IsAdmin, IsRegularUser
from datetime import timedelta, datetime, timezone as dt_timezone
import json
import os
import mimetypes
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from .funtions import compare_excel_sheets, get_sheet_names
from .funtions import get_header
from .funtions import preprocess_results
from django.http import JsonResponse, HttpResponse, FileResponse
from django.db.models import Q
from rest_framework.pagination import PageNumberPagination
from .encryption import encrypt_file, decrypt_file
from rest_framework.decorators import api_view, permission_classes
from django.db.models import Count
from django.db.models.functions import TruncDay
from datetime import datetime, timedelta
from rest_framework.permissions import IsAuthenticated
from .Connect_To_Oracle import test_connection, get_all_schemas, get_tables_in_schema, get_table_columns, compare_database_tables

class RegisterView(APIView):

    def post(self, request):

        serializer = RegisterSerializer(data=request.data)

        if serializer.is_valid():

            user = serializer.save()

            refresh = RefreshToken.for_user(user)

            return Response({

                'data': RegisterSerializer(user).data,

                'message': 'User registered successfully.',

                'token': str(refresh.access_token),

            }, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LoginView(APIView):

    def post(self, request):

        serializer = LoginSerializer(data=request.data)

        if serializer.is_valid():

            user = serializer.validated_data

            refresh = RefreshToken.for_user(user)

            return Response({

                'data': RegisterSerializer(user).data,

                'message': 'Login successful.',

                'token': str(refresh.access_token),

            })

        return Response({'message': 'Invalid credentials.'}, status=status.HTTP_401_UNAUTHORIZED)

class TokenExpiredView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        token_created_at = request.auth.payload.get('iat', None)
        if token_created_at:
            expiration_minutes = 1200
            # Use datetime.timezone.utc for proper timezone-aware comparison
            created_time = datetime.fromtimestamp(token_created_at, tz=dt_timezone.utc)
            if timezone.now() > created_time + timedelta(minutes=expiration_minutes):
                return Response({
                    'data': [],
                    'message': 'Token expired.',
                    'token': None
                }, status=status.HTTP_401_UNAUTHORIZED)

        return Response({
            'data': RegisterSerializer(request.user).data,
            'message': 'Token valid.',
            'token': str(request.auth)
        })

class UploadFileView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file = request.FILES.get('file')

        if FileUpload.objects.filter(user=request.user, file__icontains=file.name).exists():
            return Response({"message": "File already uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        file_upload = FileUpload.objects.create(
            user=request.user, file=file
        )
        
        # Get the path to the saved file
        file_path = file_upload.file.path
        
        # Encrypt the file
        try:
            encrypt_file(file_path)
        except Exception as e:
            # If encryption fails, delete the file and return an error
            file_upload.delete()
            return Response({'error': f'Failed to encrypt file: {str(e)}'}, status=500)
        
        return Response({"message": "Files uploaded and encrypted successfully.", "file_id": file_upload.id}, status=status.HTTP_201_CREATED)

    def delete(self, request, file_id):
        # Check if user is admin or has delete permission
        if request.user.role != 'admin' and not request.user.can_delete_files:
            return Response({
                "message": "You don't have permission to delete files."
            }, status=status.HTTP_403_FORBIDDEN)
        
        file_upload = FileUpload.objects.filter(id=file_id).first()
        if not file_upload:
            return Response({"message": "File not found."}, status=status.HTTP_404_NOT_FOUND)
        
        try:
            # Get the full path to the file
            file_path = file_upload.file.path
            
            # Delete the database record
            file_upload.delete()
            
            # Check if the file still exists on the filesystem and delete it
            if os.path.exists(file_path):
                os.remove(file_path)
                
            # Also check for the directory - if empty, remove it
            directory = os.path.dirname(file_path)
            if os.path.exists(directory) and not os.listdir(directory):
                os.rmdir(directory)
                
            return Response({"message": "File deleted successfully."}, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                "message": f"Error deleting file: {str(e)}",
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ListUserFilesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        files = FileUpload.objects.filter(user=request.user)
        serializer = FileUploadSerializer(files, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class ListAllFilesAdminView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        files = FileUpload.objects.all()
        serializer = FileUploadSerializer(files, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class AddComparisonView(APIView):
    permission_classes = [IsAuthenticated]  # Restrict access to authenticated users

    def post(self, request, file1_id, file2_id):
        sheet1 = request.data.get("sheet1")
        sheet2 = request.data.get("sheet2")
        range1 = request.data.get("range1", "A7:F99")
        range2 = request.data.get("range2", "A7:F99")
        column1 = request.data.get("column1")
        column2 = request.data.get("column2")

        # Improved logging for sheet names, ranges, and columns
        print(f"Received comparison request: file1_id={file1_id}, file2_id={file2_id}, sheet1={sheet1}, sheet2={sheet2}, range1={range1}, range2={range2}, column1={column1}, column2={column2}")
        print(f"Request user: {request.user.username} (ID: {request.user.id})")
        if not file1_id or not file2_id:
            return Response(
                {"error": "file1, file2, and results are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if column1 != column2:
            return Response(
                {"error": "primary key in file 1 not equal file2  primary key"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            # Fetch the files from the database
            file1 = FileUpload.objects.get(id=file1_id, user=request.user)
            file2 = FileUpload.objects.get(id=file2_id, user=request.user)

            # Ensure the files are not the same
            if file1.id == file2.id:
                return Response(
                    {"error": "file1 and file2 must be different files."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Create a list containing the primary key columns
            primary_columns = None
            if column1 and column2:
                primary_columns = [column1, column2]


            results_json = compare_excel_sheets(
                file1.file.path,
                file2.file.path,
                sheet_name=sheet1,
                range_file1=range1,
                range_file2=range2,
                PrimaryColumn=primary_columns
            )

            results_json = preprocess_results(results_json)

            # Convert Python dictionary to JSON string before saving
            comparison = Comparison.objects.create(
                file1=file1,
                file2=file2,
                results=json.dumps(results_json),
                user_id=request.user.id,
                range1=range1,
                range2=range2, 
                column1=column1,
                column2=column2,
            )

            # Serialize the created comparison
            serializer = ComparisonSerializer(comparison)
            return Response(
                {"message": "Comparison created successfully", "data": serializer.data},
                status=status.HTTP_201_CREATED,
            )
        except FileUpload.DoesNotExist:
            # Return an error if one or both files do not exist or do not belong to the user
            return Response(
                {"error": "One or both files not found or do not belong to the user."},
                status=status.HTTP_404_NOT_FOUND,
            )

class deleteComparisonView(APIView):
    permission_classes = [IsAuthenticated]  # Restrict access to authenticated users

    def delete(self, comparison_id):
        comparison = Comparison.objects.filter(id=comparison_id).first()
        if not comparison:
            return Response({"message": "File not found."}, status=status.HTTP_404_NOT_FOUND)

        comparison.delete()
        return Response({"message": "File deleted successfully."}, status=status.HTTP_200_OK)

class list_user_ComparisonView(APIView):
    permission_classes = [IsAuthenticated]  # Restrict access to authenticated users

    def get(self, request):
        comparisons = Comparison.objects.filter(user=request.user)
    
        # Optional search by filename
        search_query = request.GET.get('search', '').strip()
        if search_query:
            comparisons = comparisons.filter(
                Q(file1__file__icontains=search_query) |
                Q(file2__file__icontains=search_query)
            )

        # Optional search by comparison ID
        comparison_id = request.GET.get('comparison_id', '').strip()
        if comparison_id:
            comparisons = comparisons.filter(id=comparison_id)

        # Date filtering
        date_filter_type = request.GET.get('date_filter_type', '')
        if date_filter_type == 'range':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            if start_date and end_date:
                # Add one day to end_date for inclusive range
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                end_date = end_date_obj.strftime('%Y-%m-%d')
                comparisons = comparisons.filter(timestamp__range=[start_date, end_date])
        elif date_filter_type == 'day' and request.GET.get('filter_date'):
            filter_date = request.GET.get('filter_date')
            comparisons = comparisons.filter(timestamp__date=filter_date)
        elif date_filter_type == 'month' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET.get('filter_date'), '%Y-%m-%d')
            comparisons = comparisons.filter(
                timestamp__year=filter_date.year,
                timestamp__month=filter_date.month
            )
        elif date_filter_type == 'year' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET.get('filter_date'), '%Y-%m-%d')
            comparisons = comparisons.filter(timestamp__year=filter_date.year)

        # Paginate results
        paginator = PageNumberPagination()
        paginator.page_size = 10
        results_page = paginator.paginate_queryset(comparisons, request)
        serializer = ComparisonSerializer(results_page, many=True)
        data = serializer.data
        for i, comp in enumerate(results_page):
            data[i]['file1_name'] = comp.file1.file.name
            data[i]['file2_name'] = comp.file2.file.name
            data[i]['user'] = comp.user.username  # Add user field
        return Response({
            'results': data,
            'current_page': paginator.page.number,
            'total_pages': paginator.page.paginator.num_pages
        }, status=status.HTTP_200_OK)

class GetSheetNamesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, file_id):
        try:
            file_obj = FileUpload.objects.get(id=file_id, user=request.user)
            names = get_sheet_names(file_obj.file.path)
            return JsonResponse({"sheet_names": names}, status=200)
        except FileUpload.DoesNotExist:
            return JsonResponse({"error": "File not found or not authorized."}, status=404)

class GetFileHeaderView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, file_id):
        # Get range from request data with default value as fallback
        range_value = request.data.get('range', 'A7:F99')
        
        # Retrieve the file path from FileUpload using file_id
        file_upload = FileUpload.objects.filter(id=file_id, user=request.user).first()
        if not file_upload:
            return Response({"error": "File not found or not yours."}, status=status.HTTP_404_NOT_FOUND)
        columns = get_header(file_upload.file.path, range=range_value)
        print(f"Retrieved columns: {columns} for file_id: {file_id}, range: {range_value}")
        return Response({"columns": columns}, status=status.HTTP_200_OK)

class ListAllComparisonView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]  # Admin only

    def get(self, request):
        # Get all comparisons with related data
        comparisons = Comparison.objects.all().select_related('file1', 'file2', 'user')

        # Apply filters if provided
        search_query = request.GET.get('search', '').strip()
        if search_query:
            comparisons = comparisons.filter(
                Q(file1__file__icontains=search_query) |
                Q(file2__file__icontains=search_query)
            )

        comparison_id = request.GET.get('comparison_id', '').strip()
        if comparison_id:
            comparisons = comparisons.filter(id=comparison_id)

        # Date filtering
        date_filter_type = request.GET.get('date_filter_type', '')
        if date_filter_type == 'range':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            if start_date and end_date:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                comparisons = comparisons.filter(timestamp__range=[start_date, end_date_obj.strftime('%Y-%m-%d')])
        elif date_filter_type == 'day' and request.GET.get('filter_date'):
            comparisons = comparisons.filter(timestamp__date=request.GET['filter_date'])
        elif date_filter_type == 'month' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET['filter_date'], '%Y-%m-%d')
            comparisons = comparisons.filter(timestamp__year=filter_date.year, timestamp__month=filter_date.month)
        elif date_filter_type == 'year' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET['filter_date'], '%Y-%m-%d')
            comparisons = comparisons.filter(timestamp__year=filter_date.year)

        # Order the query to mix users' data - sort by user ID first, then timestamp
        comparisons = comparisons.order_by('user__id', '-timestamp')

        # Larger page size to show more results
        paginator = PageNumberPagination()
        paginator.page_size = 10
        results_page = paginator.paginate_queryset(comparisons, request)
        data = []
        for comp in results_page:
            data.append({
                'id': comp.id,
                'file1_name': comp.file1.file.name,
                'file2_name': comp.file2.file.name,
                'user': comp.user.username,
                'timestamp': comp.timestamp
            })
        return paginator.get_paginated_response(data)

class ComparisonDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, comparison_id):
        # For admin users, allow access to any comparison
        if request.user.role == 'admin':
            comparison = Comparison.objects.filter(id=comparison_id).first()
        else:
            # For regular users, only allow access to their own comparisons
            comparison = Comparison.objects.filter(id=comparison_id, user=request.user).first()

        if not comparison:
            return Response({"error": "Comparison not found or you don't have permission to view it"}, 
                           status=status.HTTP_404_NOT_FOUND)

        # Serialize the comparison with all details including results
        data = {
            'id': comparison.id,
            'file1_name': comparison.file1.file.name,
            'file2_name': comparison.file2.file.name,
            'user': comparison.user.username,
            'timestamp': comparison.timestamp,
            'results': comparison.results  # Include the full results field
        }
        return Response(data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_analytics(request):
    """ 
    Returns analytics data for the dashboard, including user and admin stats.
    """
    user = request.user
    is_admin = (getattr(user, 'role', None) == 'admin')

    # User-specific stats
    user_files = FileUpload.objects.filter(user=user)
    user_file_count = user_files.count()
    user_comparisons = Comparison.objects.filter(user=user)
    user_comparison_count = user_comparisons.count()

    # Simple file-types breakdown for user
    user_file_types = {}
    for f in user_files:
        ext = f.file.name.split('.')[-1].lower()
        user_file_types[ext] = user_file_types.get(ext, 0) + 1

    # Recent comparisons
    recent_comps = user_comparisons.order_by('-timestamp')[:5]

    # Activity by day (last 30 days)
    thirty_days_ago = datetime.now() - timedelta(days=30)
    daily_activity = user_comparisons.filter(timestamp__gte=thirty_days_ago).annotate(
        day=TruncDay('timestamp')
    ).values('day').annotate(count=Count('id')).order_by('day')
    activity_data = {
        item['day'].strftime('%Y-%m-%d'): item['count']
        for item in daily_activity
    }

    data = {
        "user_stats": {
            "total_files": user_file_count,
            "total_comparisons": user_comparison_count,
            "file_types": user_file_types,
            "recent_comparisons": [
                {
                    "file1": comp.file1.file.name.split('/')[-1],
                    "file2": comp.file2.file.name.split('/')[-1],
                    "timestamp": comp.timestamp,
                }
                for comp in recent_comps
            ],
            "activity_by_day": activity_data
        }
    }

    # Admin stats
    if is_admin:
        total_users = User.objects.count()
        total_files = FileUpload.objects.count()
        total_comparisons = Comparison.objects.count()

        top_users_query = Comparison.objects.values('user__username').annotate(
            comparison_count=Count('id')
        ).order_by('-comparison_count')[:5]

        system_file_types = {}
        for f in FileUpload.objects.all():
            ext = f.file.name.split('.')[-1].lower()
            system_file_types[ext] = system_file_types.get(ext, 0) + 1

        recent_sys_activity = Comparison.objects.order_by('-timestamp')[:10]
        sys_daily_activity = Comparison.objects.filter(timestamp__gte=thirty_days_ago).annotate(
            day=TruncDay('timestamp')
        ).values('day').annotate(count=Count('id')).order_by('day')
        sys_activity_data = {
            item['day'].strftime('%Y-%m-%d'): item['count']
            for item in sys_daily_activity
        }

        data["admin_stats"] = {
            "total_users": total_users,
            "total_files": total_files,
            "total_comparisons": total_comparisons,
            "top_users": [
                {
                    "username": item['user__username'],
                    "comparison_count": item['comparison_count']
                }
                for item in top_users_query
            ],
            "system_file_types": system_file_types,
            "recent_system_activity": [
                {
                    "user": comp.user.username,
                    "file1": comp.file1.file.name.split('/')[-1],
                    "file2": comp.file2.file.name.split('/')[-1],
                    "timestamp": comp.timestamp,
                }
                for comp in recent_sys_activity
            ],
            "system_activity_by_day": sys_activity_data
        }

    return Response(data)

class ListUsersView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]  # Admin only
    
    def get(self, request):
        users = User.objects.exclude(id=request.user.id)  # Exclude current admin
        data = []
        for user in users:
            data.append({
                'id': user.id,
                'username': user.username,
                'role': user.role,
                'can_delete_files': user.can_delete_files
            })
        return Response(data)

class UpdateUserPermissionView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]  # Admin only
    
    def put(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            can_delete_files = request.data.get('can_delete_files', False)
            
            # Update the permission
            user.can_delete_files = can_delete_files
            user.save()
            
            return Response({
                "message": f"Permissions updated for user {user.username}",
                "can_delete_files": user.can_delete_files
            })
        except User.DoesNotExist:
            return Response({"message": "User not found."}, status=status.HTTP_404_NOT_FOUND)

class CreateDatabaseConnectionView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        # Add user to the data
        data = request.data.copy()
        
        # Create the serializer with the data including the user
        serializer = DatabaseConnectionSerializer(data=data)
        
        if serializer.is_valid():
            # Save with the current user
            connection = serializer.save(user=request.user)
            
            return Response({
                'message': 'Database connection created successfully.',
                'data': DatabaseConnectionSerializer(connection).data
            }, status=status.HTTP_201_CREATED)
            
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ListUserDatabaseConnectionsView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        connections = Database_Connection.objects.filter(user=request.user)
        serializer = DatabaseConnectionSerializer(connections, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class ListAllDatabaseConnectionsView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]  # Admin only
    
    def get(self, request):
        connections = Database_Connection.objects.all()
        serializer = DatabaseConnectionSerializer(connections, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class UpdateDatabaseConnectionView(APIView):
    permission_classes = [IsAuthenticated]
    
    def put(self, request, connection_id):
        try:
            connection = Database_Connection.objects.get(id=connection_id, user=request.user)
            
            # Update the connection details
            serializer = DatabaseConnectionSerializer(connection, data=request.data, partial=True)
            
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'message': 'Database connection updated successfully.',
                    'data': serializer.data
                }, status=status.HTTP_200_OK)
                
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
        except Database_Connection.DoesNotExist:
            return Response(
                {'message': 'Database connection not found or you do not have permission to modify it.'},
                status=status.HTTP_404_NOT_FOUND
            )

class DeleteDatabaseConnectionView(APIView):
    permission_classes = [IsAuthenticated]
    
    def delete(self, request, connection_id):
        try:
            connection = Database_Connection.objects.get(id=connection_id, user=request.user)
            connection.delete()
            
            return Response(
                {'message': 'Database connection deleted successfully.'},
                status=status.HTTP_200_OK
            )
            
        except Database_Connection.DoesNotExist:
            return Response(
                {'message': 'Database connection not found or you do not have permission to delete it.'},
                status=status.HTTP_404_NOT_FOUND
            )

class TestDatabaseConnectionView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request, connection_id=None):
        try:
            # If connection_id is provided, use stored credentials, otherwise use provided ones
            if connection_id:
                # Get the connection details
                connection = Database_Connection.objects.get(id=connection_id, user=request.user)
                username = connection.username
                password = connection.password
                dsn = connection.DSN
                schema = getattr(connection, 'SCHEMA', None)
            else:
                # Use credentials from request body
                username = request.data.get('username')
                password = request.data.get('password')
                dsn = request.data.get('DSN')
                schema = request.data.get('SCHEMA', None)
                
                if not username or not password or not dsn:
                    return Response({
                        'message': 'Username, password, and DSN are required.',
                        'status': 'failed'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Test the database connection using actual function from Connect_To_Oracle.py
            try:
                result = test_connection(username, password, dsn, schema)
                
                if result["status"] == "connected":
                    return Response({
                        'message': result["message"],
                        'status': 'connected',
                        'schema_valid': result.get("schema_valid", True)
                    }, status=status.HTTP_200_OK)
                else:
                    return Response({
                        'message': result["message"],
                        'status': 'failed'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
            except Exception as e:
                # If connection test fails, return error
                return Response({
                    'message': f'Failed to connect to the database: {str(e)}',
                    'status': 'failed'
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Database_Connection.DoesNotExist:
            return Response({
                'message': 'Database connection not found or you do not have permission to access it.',
                'status': 'failed'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'message': f'Error: {str(e)}',
                'status': 'failed'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class TestDatabaseConnectionByIdView(APIView):
    """Test an existing database connection by ID"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request, connection_id):
        # Reuse the implementation from TestDatabaseConnectionView
        return TestDatabaseConnectionView().post(request, connection_id)

class GetAvailableSchemasView(APIView):
    """Returns all available schemas for a database connection"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request, connection_id=None):
        try:
            # Get connection parameters either from stored connection or request body
            if connection_id:
                # Check if user is admin - if so, allow access to any connection
                if request.user.role == 'admin':
                    connection = Database_Connection.objects.get(id=connection_id)
                else:
                    # Regular users can only access their own connections
                    connection = Database_Connection.objects.get(id=connection_id, user=request.user)
                
                username = connection.username
                password = connection.password
                dsn = connection.DSN
            else:
                # Use credentials from request body
                username = request.data.get('username')
                password = request.data.get('password')
                dsn = request.data.get('DSN')
                
                if not username or not password or not dsn:
                    return Response({
                        'message': 'Username, password, and DSN are required.',
                        'status': 'error'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get all available schemas
            try:
                schemas = get_all_schemas(username, password, dsn)
                return Response({
                    'status': 'success',
                    'schemas': schemas
                }, status=status.HTTP_200_OK)
                
            except Exception as e:
                return Response({
                    'message': f'Failed to retrieve schemas: {str(e)}',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Database_Connection.DoesNotExist:
            return Response({
                'message': 'Database connection not found or you do not have permission to access it.',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'message': f'Error: {str(e)}',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class GetSchemaTablesView(APIView):
    """Returns all tables within a specific schema"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request, connection_id=None):
        try:
            # Get schema name from the request
            schema_name = request.data.get('schema')
            if not schema_name:
                return Response({
                    'message': 'Schema name is required.',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get connection parameters either from stored connection or request body
            if connection_id:
                # Check if user is admin - if so, allow access to any connection
                if request.user.role == 'admin':
                    connection = Database_Connection.objects.get(id=connection_id)
                else:
                    # Regular users can only access their own connections
                    connection = Database_Connection.objects.get(id=connection_id, user=request.user)
                    
                username = connection.username
                password = connection.password
                dsn = connection.DSN
            else:
                # Use credentials from request body
                username = request.data.get('username')
                password = request.data.get('password')
                dsn = request.data.get('DSN')
                
                if not username or not password or not dsn:
                    return Response({
                        'message': 'Username, password, and DSN are required.',
                        'status': 'error'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get all tables for the specified schema
            try:
                tables = get_tables_in_schema(username, password, dsn, schema_name)
                return Response({
                    'status': 'success',
                    'schema': schema_name,
                    'tables': tables
                }, status=status.HTTP_200_OK)
                
            except Exception as e:
                return Response({
                    'message': f'Failed to retrieve tables for schema {schema_name}: {str(e)}',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Database_Connection.DoesNotExist:
            return Response({
                'message': 'Database connection not found or you do not have permission to access it.',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'message': f'Error: {str(e)}',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class GetTableColumnsView(APIView):
    """Returns all columns for a specific table in a schema"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request, connection_id=None):
        try:
            # Get schema and table name from the request
            schema_name = request.data.get('schema')
            table_name = request.data.get('table')
            
            if not schema_name or not table_name:
                return Response({
                    'message': 'Schema name and table name are required.',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get connection parameters from stored connection
            if connection_id:
                # Check if user is admin - if so, allow access to any connection
                if request.user.role == 'admin':
                    connection = Database_Connection.objects.get(id=connection_id)
                else:
                    # Regular users can only access their own connections
                    connection = Database_Connection.objects.get(id=connection_id, user=request.user)
                    
                username = connection.username
                password = connection.password
                dsn = connection.DSN
            else:
                # Use credentials from request body
                username = request.data.get('username')
                password = request.data.get('password')
                dsn = request.data.get('DSN')
                
                if not username or not password or not dsn:
                    return Response({
                        'message': 'Username, password, and DSN are required.',
                        'status': 'error'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get columns for the specified table
            try:
                columns_result = get_table_columns(username, password, dsn, schema_name, table_name)
                
                if 'error' in columns_result:
                    return Response({
                        'message': columns_result['error'],
                        'status': 'error'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                return Response({
                    'status': 'success',
                    'columns': columns_result['columns']
                }, status=status.HTTP_200_OK)
                
            except Exception as e:
                return Response({
                    'message': f'Failed to retrieve columns for table {schema_name}.{table_name}: {str(e)}',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Database_Connection.DoesNotExist:
            return Response({
                'message': 'Database connection not found or you do not have permission to access it.',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'message': f'Error: {str(e)}',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class CompareDatabaseTablesView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        # Get connection parameters
        connection_id1 = request.data.get('connection1')
        connection_id2 = request.data.get('connection2')
        schema1 = request.data.get('schema1')
        schema2 = request.data.get('schema2')
        table1 = request.data.get('table1')
        table2 = request.data.get('table2')
        primary_key1 = request.data.get('column1')
        primary_key2 = request.data.get('column2')
        
        # Validate required parameters
        if not all([connection_id1, connection_id2, schema1, schema2, table1, table2, primary_key1, primary_key2]):
            return Response({
                'message': 'All parameters are required: connection1, connection2, schema1, schema2, table1, table2, column1, column2',
                'status': 'error'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Get connection details for both connections
            # connection1 = Database_Connection.objects.get(id=connection_id1, user=request.user)
            # connection2 = Database_Connection.objects.get(id=connection_id2, user=request.user)
            connection1 = Database_Connection.objects.get(id=connection_id1)
            connection2 = Database_Connection.objects.get(id=connection_id2)
            
            # Get credentials for both connections
            username1 = connection1.username
            password1 = connection1.password
            dsn1 = connection1.DSN
            
            username2 = connection2.username
            password2 = connection2.password
            dsn2 = connection2.DSN
            
            # Compare tables using the updated function with both sets of credentials
            results = compare_database_tables(
                username1, password1, dsn1, schema1, table1, primary_key1,
                username2, password2, dsn2, schema2, table2, primary_key2
            )
            
            # Save comparison if requested
            save_comparison = request.data.get('save_comparison', False)
            comparison_id = None
            
            if save_comparison:
                # Create a new database comparison record
                db_comparison = Database_Comparison.objects.create(
                    user=request.user,
                    connection1=connection1,
                    connection2=connection2,
                    schema1=schema1,
                    schema2=schema2,
                    table1=table1,
                    table2=table2,
                    primary_key1=primary_key1,
                    primary_key2=primary_key2,
                    results=results
                )
                comparison_id = db_comparison.id
            
            return Response({
                'status': 'success',
                'message': 'Tables compared successfully',
                'results': results,
                'saved': save_comparison,
                'comparison_id': comparison_id
            }, status=status.HTTP_200_OK)
            
        except Database_Connection.DoesNotExist:
            return Response({
                'message': 'One or both database connections not found or unauthorized access',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'message': f'Error comparing tables: {str(e)}',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ListUserDatabaseComparisonsView(APIView):
    """Returns database comparisons with pagination and filtering support"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Start with all comparisons for the user or all comparisons for admin
        if request.user.role == 'admin' and request.GET.get('all', '').lower() == 'true':
            comparisons = Database_Comparison.objects.all()
        else:
            comparisons = Database_Comparison.objects.filter(user=request.user)
        
        # Apply filters
        search_query = request.GET.get('search', '').strip()
        if search_query:
            comparisons = comparisons.filter(
                Q(schema1__icontains=search_query) | 
                Q(schema2__icontains=search_query) |
                Q(table1__icontains=search_query) |
                Q(table2__icontains=search_query)
            )
        
        # Date filtering
        date_filter_type = request.GET.get('date_filter_type', '')
        if date_filter_type == 'range':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            if start_date and end_date:
                # Add one day to end_date for inclusive range
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                end_date = end_date_obj.strftime('%Y-%m-%d')
                comparisons = comparisons.filter(timestamp__range=[start_date, end_date])
        elif date_filter_type == 'day' and request.GET.get('filter_date'):
            filter_date = request.GET.get('filter_date')
            comparisons = comparisons.filter(timestamp__date=filter_date)
        elif date_filter_type == 'month' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET.get('filter_date'), '%Y-%m-%d')
            comparisons = comparisons.filter(
                timestamp__year=filter_date.year,
                timestamp__month=filter_date.month
            )
        elif date_filter_type == 'year' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET.get('filter_date'), '%Y-%m-%d')
            comparisons = comparisons.filter(timestamp__year=filter_date.year)
        
        # Order by timestamp (most recent first)
        comparisons = comparisons.order_by('-timestamp')
        
        # Paginate results
        paginator = PageNumberPagination()
        paginator.page_size = 10
        page = paginator.paginate_queryset(comparisons, request)
        
        if page is not None:
            # Serialize the paginated results
            data = []
            for comp in page:
                # Get usernames for display
                try:
                    connection1_username = comp.connection1.username
                    connection2_username = comp.connection2.username
                except:
                    connection1_username = "Unknown"
                    connection2_username = "Unknown"
                
                data.append({
                    'id': comp.id,
                    'source': f"{comp.schema1}.{comp.table1}",
                    'target': f"{comp.schema2}.{comp.table2}",
                    'connection1_name': f"{connection1_username}@{comp.connection1.DSN}",
                    'connection2_name': f"{connection2_username}@{comp.connection2.DSN}",
                    'timestamp': comp.timestamp,
                    'summary': {
                        'total_columns_table1': comp.results.get('summary', {}).get('total_columns_table1', 0),
                        'total_columns_table2': comp.results.get('summary', {}).get('total_columns_table2', 0),
                        'added_rows': comp.results.get('summary', {}).get('added_rows', 0),
                        'removed_rows': comp.results.get('summary', {}).get('removed_rows', 0),
                        'changed_values': comp.results.get('summary', {}).get('changed_values', 0),
                    }
                })
            
            return Response({
                'results': data,
                'current_page': paginator.page.number,
                'total_pages': paginator.page.paginator.num_pages
            })
        
        # If pagination is not needed (fallback)
        data = []
        for comp in comparisons:
            data.append({
                'id': comp.id,
                'source': f"{comp.schema1}.{comp.table1}",
                'target': f"{comp.schema2}.{comp.table2}",
                'timestamp': comp.timestamp
            })
            
        return Response(data)

class DeleteDatabaseComparisonView(APIView):
    """Deletes a database comparison by ID"""
    permission_classes = [IsAuthenticated]
    
    def delete(self, request, comparison_id):
        try:
            # For admin users, allow deleting any comparison
            if request.user.role == 'admin':
                comparison = Database_Comparison.objects.get(id=comparison_id)
            else:
                # For regular users, only allow deleting their own comparisons
                comparison = Database_Comparison.objects.get(id=comparison_id, user=request.user)
            
            comparison.delete()
            return Response({
                'message': 'Database comparison deleted successfully',
                'status': 'success'
            }, status=status.HTTP_200_OK)
        except Database_Comparison.DoesNotExist:
            return Response({
                'message': 'Comparison not found or you do not have permission to delete it',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)

class GetDatabaseComparisonView(APIView):
    """Returns details of a specific database comparison"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, comparison_id):
        try:
            # For admin, allow access to any comparison
            if request.user.role == 'admin':
                comparison = Database_Comparison.objects.get(id=comparison_id)
            else:
                # For regular users, only allow access to their own comparisons
                comparison = Database_Comparison.objects.get(id=comparison_id, user=request.user)
                
            return Response({
                'id': comparison.id,
                'connection1': comparison.connection1.id,
                'connection2': comparison.connection2.id,
                'schema1': comparison.schema1,
                'schema2': comparison.schema2,
                'table1': comparison.table1,
                'table2': comparison.table2,
                'primary_key1': comparison.primary_key1,
                'primary_key2': comparison.primary_key2,
                'timestamp': comparison.timestamp,
                'results': comparison.results,
            }, status=status.HTTP_200_OK)
        except Database_Comparison.DoesNotExist:
            return Response({
                'message': 'Comparison not found or you do not have permission to view it',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)




############
def column_letter_to_index(col_letter):
    """Convert column letter (A, B, C, etc.) to 0-based index"""
    result = 0
    for char in col_letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1
def get_excel_column_name(column_index):
    """Convert 0-based column index to Excel column letter(s)"""
    result = ""
    while column_index >= 0:
        result = chr(column_index % 26 + ord('A')) + result
        column_index = column_index // 26 - 1
        if column_index < 0:
            break
    return result
def get_next_column_letter(col_letter):
    """Get the next column letter (A->B, Z->AA, etc.)"""
    col_letter = col_letter.upper()
    
    # Convert to list for easier manipulation
    letters = list(col_letter)
    
    # Start from the rightmost character
    i = len(letters) - 1
    
    while i >= 0:
        if letters[i] != 'Z':
            # Simple case: increment the letter
            letters[i] = chr(ord(letters[i]) + 1)
            break
        else:
            # This letter is Z, so it becomes A and we carry over
            letters[i] = 'A'
            if i == 0:
                # We've carried over past the first letter, so add a new A at the beginning
                letters.insert(0, 'A')
                break
            # Continue to the next letter to the left
            i -= 1
    
    return ''.join(letters)
def safe_assign_to_dataframe(df, row_index, column_index, value, column_name=None):
    """
    Safely assign a value to df.iloc[row_index, column_index].
    Expands the DataFrame if necessary.
    """
    # Ensure we have enough rows
    if row_index >= len(df):
        # Add missing rows
        missing_rows = row_index - len(df) + 1
        for _ in range(missing_rows):
            df.loc[len(df)] = [None] * len(df.columns)
    
    # Ensure we have enough columns
    current_columns = len(df.columns)
    if column_index >= current_columns:
        # Add missing columns
        for i in range(current_columns, column_index + 1):
            if column_name and i == column_index:
                col_name = column_name
            else:
                col_name = get_excel_column_name(i)
            df[col_name] = None
    
    # Now safely assign the value
    df.iloc[row_index, column_index] = value
    return df

import pandas as pd
import io
import re
import os
import json

# Add these imports at the top if not already present
import pandas as pd
import io
import re
import os
import json
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class DownloadComparisonExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, comparison_id, file_number=1):
        """
        Download the Excel file from a comparison.
        file_number: 1 for first file, 2 for second file
        """
        try:
            # For admin, allow access to any comparison
            if request.user.role == 'admin':
                comparison = Comparison.objects.get(id=comparison_id)
            else:
                # For regular users, only allow access to their own comparisons
                comparison = Comparison.objects.get(id=comparison_id, user=request.user)
                
            # Determine which file to download
            if file_number == 1:
                file_obj = comparison.file1
                file_label = "source"
            elif file_number == 2:
                file_obj = comparison.file2
                file_label = "target"
            else:
                return Response({
                    'message': 'Invalid file number. Use 1 for source file or 2 for target file.',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get the file path and comparison data
            range1 = comparison.range1
            range2 = comparison.range2
            result = comparison.results
            
            if isinstance(result, str):
                result_data = json.loads(result)
            else:
                result_data = result
                
            print(f"Range1: {range1}, Range2: {range2}")
            print(f"Comparison Results: {result}")
            
            file_path = file_obj.file.path
            
            # Check if file exists
            if not os.path.exists(file_path):
                return Response({
                    'message': f'File not found: {file_label} file for comparison {comparison_id}',
                    'status': 'error'
                }, status=status.HTTP_404_NOT_FOUND)
                
            # Decrypt the file temporarily for download
            try:
                # Check the file size before decryption
                file_size = os.path.getsize(file_path)
                print(f"Original file size: {file_size} bytes")
                
                # First, try to read the file directly to see if it's already decrypted
                with open(file_path, 'rb') as f:
                    first_bytes = f.read(10)
                    print(f"First 10 bytes of file: {first_bytes}")
                
                # Check if the file is already in Excel format (not encrypted)
                if first_bytes.startswith(b'PK') or first_bytes.startswith(b'\xd0\xcf\x11\xe0'):
                    print("File appears to be unencrypted Excel format")
                    # Read the entire file as-is
                    with open(file_path, 'rb') as f:
                        decrypted_content = f.read()
                    print(f"Direct file read size: {len(decrypted_content)} bytes")
                else:
                    print("File appears to be encrypted, attempting decryption...")
                    # Try to decrypt the file
                    try:
                        decrypted_result = decrypt_file(file_path)
                        print(f"Decrypt function returned: type={type(decrypted_result)}, length={len(str(decrypted_result))}")
                        
                        # Check if decrypt_file returned a path (string) instead of content
                        if isinstance(decrypted_result, str):
                            if os.path.exists(decrypted_result):
                                print(f"Decrypt function returned a file path: {decrypted_result}")
                                print("Reading the decrypted file content...")
                                with open(decrypted_result, 'rb') as f:
                                    decrypted_content = f.read()
                                print(f"File content read successfully: {len(decrypted_content)} bytes")
                            else:
                                print("Decrypt function returned a string, but it's not a valid path")
                                print(f"First 100 characters: {decrypted_result[:100]}")
                                # Try to encode the string as bytes
                                decrypted_content = decrypted_result.encode('latin-1')
                        else:
                            # Decrypt function returned bytes directly
                            decrypted_content = decrypted_result
                            
                        print(f"Final decrypted content size: {len(decrypted_content)} bytes")
                        print(f"Final decrypted content type: {type(decrypted_content)}")
                        
                    except Exception as decrypt_error:
                        print(f"Decryption failed: {str(decrypt_error)}")
                        return Response({
                            'message': f'Error decrypting file: {str(decrypt_error)}',
                            'status': 'error'
                        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
                # Ensure decrypted_content is bytes
                if isinstance(decrypted_content, str):
                    print("Converting string to bytes using latin-1 encoding")
                    try:
                        decrypted_content = decrypted_content.encode('latin-1')
                    except UnicodeEncodeError:
                        print("Failed to encode with latin-1, trying utf-8")
                        try:
                            decrypted_content = decrypted_content.encode('utf-8')
                        except UnicodeEncodeError:
                            print("Both encoding methods failed, file might be corrupted")
                            return Response({
                                'message': 'File encoding error - file may be corrupted',
                                'status': 'error'
                            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
                # Verify we have bytes now
                if not isinstance(decrypted_content, bytes):
                    return Response({
                        'message': f'Error: Decrypted content is {type(decrypted_content)}, expected bytes',
                        'status': 'error'
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                    
                # Verify the decrypted content starts with valid Excel magic bytes
                if decrypted_content.startswith(b'PK'):
                    print("File appears to be a valid zip-based format (Excel .xlsx)")
                elif decrypted_content.startswith(b'\xd0\xcf\x11\xe0'):
                    print("File appears to be OLE format (Excel .xls)")
                else:
                    print(f"Warning: File doesn't start with expected Excel headers. First 10 bytes: {decrypted_content[:10]}")
                    print("File might still be valid - proceeding with download")
                    
                # Modify the existing Excel file structure
                try:
                    # Create a BytesIO object from the decrypted content
                    excel_buffer = io.BytesIO(decrypted_content)
                    
                    # Load the workbook using openpyxl to preserve original structure
                    workbook = load_workbook(excel_buffer)
                    
                    # Get the first worksheet (or specify sheet name if needed)
                    worksheet = workbook.active  # or workbook['SheetName']
                    
                    # Parse the range
                    rangee = range2.split(':')
                    Range1 = rangee[0].strip().upper()
                    Range2 = rangee[1].strip().upper() if len(rangee) > 1 else None

                    match = re.match(r'([A-Z]+)(\d+)', Range1)
                    match2 = re.match(r'([A-Z]+)(\d+)', Range2)
                    
                    if match and match2:
                        col_letter = match.group(1)
                        row_number = int(match.group(2))
                        col_letter2 = match2.group(1)
                        row_number2 = int(match2.group(2))
                        next_column = get_next_column_letter(col_letter2)
                        
                        print(f"Current column: {col_letter2}, Next column: {next_column}")

                        col_letter_index = column_letter_to_index(col_letter)
                        next_column_index = column_letter_to_index(next_column)
                        
                        # Get comparison results
                        added_row = result_data.get("rows_added", [])
                        removed_row = result_data.get("rows_removed", [])
                        changed_values = result_data.get("value_diff", [])
                        column= comparison.column1
                        
                        # Create a set of rows to mark as unmatched for efficiency
                        unmatched_rows = set()
                        workbook2 = load_workbook(excel_buffer)
                        worksheet2 = workbook2.active
                        store_matched_rows = []

                        for row_num in range(row_number+2, row_number2): # +1 to skip header row
                            cell_value = worksheet2.cell(row=row_num, column=col_letter_index+1).value
                            if cell_value is None or cell_value == '':
                                continue
                            store_matched_rows.append(row_num)
                            print(cell_value) 
                        
                        # Process added rows
                        for item in added_row:                            
                              if column in item and column in item and item[column] is not None:
                                unmatched_rows.add(item['_OriginalRow'])
                        

                        # Process changed values
                        for item in changed_values:
                            if file_number == 1 and 'excel_row_file1' in item:
                                unmatched_rows.add(item['excel_row_file1'])
                            elif file_number == 2 and 'excel_row_file2' in item:
                                unmatched_rows.add(item['excel_row_file2'])
                        
                        # Add header for the new column if it doesn't exist
                        header_cell = worksheet.cell(row=row_number, column=next_column_index + 1)
                        if header_cell.value is None or header_cell.value == '':
                            header_cell.value = "Comparison Status"
                            header_cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Yellow

                        

                        total_matched = 0
                        total_unmatched = len(unmatched_rows)   
                        # Mark unmatched rows in the new column
                        for row_target in range(row_number + 2, row_number2 ):  # +1 to skip header row
                            cell = worksheet.cell(row=row_target, column=next_column_index + 1)
                            
                            if row_target in unmatched_rows:
                                cell.value = "unmatched"
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
                                print(f"Row {row_target} in column {next_column} marked as unmatched")

                            elif row_target in store_matched_rows:
                                cell.value = "match"
                                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                                total_matched += 1
                                print(f"Row {row_target} in column {next_column} marked as match")

                        total_rows_with_data = len(store_matched_rows)
                        matched_percentage = (total_matched / total_rows_with_data * 100) if total_rows_with_data > 0 else 0
                        unmatched_percentage = (total_unmatched / total_rows_with_data * 100) if total_rows_with_data > 0 else 0

                        cell=worksheet.cell(row=row_number , column=next_column_index + 2)
                        cell.value = "Total matched"
                        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Medium gray for header



                        cell = worksheet.cell(row=row_number + 3, column=next_column_index + 2)
                        cell.value = f"{total_matched} ({matched_percentage:.1f}%)"
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                          # Light gray

                        cell = worksheet.cell(row=row_number, column=next_column_index + 3)
                        cell.value = "Total unmatched"
                        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

                        cell = worksheet.cell(row=row_number + 3, column=next_column_index + 3)
                        cell.value = f"{total_unmatched} ({unmatched_percentage:.1f}%)"
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow

                        comparison_status_col = get_excel_column_name(next_column_index)
                        worksheet.column_dimensions[comparison_status_col].width = 20

                        comparison_status_col = get_excel_column_name(next_column_index+1)
                        worksheet.column_dimensions[comparison_status_col].width = 20

                        comparison_status_col = get_excel_column_name(next_column_index+2)
                        worksheet.column_dimensions[comparison_status_col].width = 20
                            
                    
                    # Save the modified workbook to a new BytesIO buffer
                    modified_excel_buffer = io.BytesIO()
                    workbook.save(modified_excel_buffer)
                    
                    # Get the modified content
                    modified_excel_buffer.seek(0)
                    final_content = modified_excel_buffer.getvalue()
                    
                    print(f"Modified Excel content size: {len(final_content)} bytes")
                    
                except Exception as excel_modify_error:
                    print(f"Error modifying Excel file: {str(excel_modify_error)}")
                    # Fallback to original content if modification fails
                    final_content = decrypted_content
                
                # Get the original filename without path
                original_filename = os.path.basename(file_obj.file.name)
                
                # Ensure proper Excel content type
                if original_filename.lower().endswith('.xlsx'):
                    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                elif original_filename.lower().endswith('.xls'):
                    content_type = 'application/vnd.ms-excel'
                elif original_filename.lower().endswith('.csv'):
                    content_type = 'text/csv'
                else:
                    # Default to Excel format
                    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    
                # Create the response with the modified content
                response = HttpResponse(final_content, content_type=content_type)
                response['Content-Disposition'] = f'attachment; filename="{original_filename}"'
                response['Content-Length'] = len(final_content)
                response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                response['Pragma'] = 'no-cache'
                response['Expires'] = '0'
                
                print(f"Sending response with {len(final_content)} bytes, content-type: {content_type}")
                return response
                
            except Exception as e:
                return Response({
                    'message': f'Error processing file: {str(e)}',
                    'status': 'error'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Comparison.DoesNotExist:
            return Response({
                'message': 'Comparison not found or you do not have permission to access it',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'message': f'Error: {str(e)}',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class AnalyzeExcelTableView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, comparison_id, file_number=1):
        """
        Analyze Excel table structure and return the first empty cell after the table.
        file_number: 1 for first file, 2 for second file
        """
        try:
            # For admin, allow access to any comparison
            if request.user.role == 'admin':
                comparison = Comparison.objects.get(id=comparison_id)
            else:
                # For regular users, only allow access to their own comparisons
                comparison = Comparison.objects.get(id=comparison_id, user=request.user)
            
            # Determine which file to analyze
            if file_number == 1:
                file_obj = comparison.file1
                file_label = "source"
            elif file_number == 2:
                file_obj = comparison.file2
                file_label = "target"
            else:
                return Response({
                    'message': 'Invalid file number. Use 1 for source file or 2 for target file.',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get the file path
            file_path = file_obj.file.path
            
            # Check if file exists
            if not os.path.exists(file_path):
                return Response({
                    'message': f'File not found: {file_label} file for comparison {comparison_id}',
                    'status': 'error'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Decrypt and analyze the file
            try:
                # First, check if file is already decrypted
                with open(file_path, 'rb') as f:
                    first_bytes = f.read(10)
                
                # Get file content
                if first_bytes.startswith(b'PK') or first_bytes.startswith(b'\xd0\xcf\x11\xe0'):
                    # File is not encrypted
                    with open(file_path, 'rb') as f:
                        file_content = f.read()
                else:
                    # File is encrypted, decrypt it
                    decrypted_result = decrypt_file(file_path)
                    if isinstance(decrypted_result, str) and os.path.exists(decrypted_result):
                        with open(decrypted_result, 'rb') as f:
                            file_content = f.read()
                    else:
                        file_content = decrypted_result if isinstance(decrypted_result, bytes) else decrypted_result.encode('latin-1')
                
                # Analyze the Excel content
                import pandas as pd
                import io
                
                excel_buffer = io.BytesIO(file_content)
                
                # Read all sheets to get their names
                excel_file = pd.ExcelFile(excel_buffer)
                sheet_names = excel_file.sheet_names
                
                analysis_results = {}
                
                # Analyze each sheet
                for sheet_name in sheet_names:
                    try:
                        df = pd.read_excel(excel_buffer, sheet_name=sheet_name, header=None)
                          # Find the last row with data
                        non_empty_rows = df.dropna(how='all')
                        if not non_empty_rows.empty:
                            last_row_with_data = non_empty_rows.index[-1]
                            first_empty_row = last_row_with_data + 2  # +1 for 1-indexed, +1 for next row
                            
                            # Find the last column with data
                            non_empty_cols = df.dropna(how='all', axis=1)
                            if not non_empty_cols.empty:
                                last_col_with_data = non_empty_cols.columns[-1]
                                # Convert column number to letter
                                first_empty_col_letter = chr(65 + last_col_with_data + 1) if last_col_with_data < 25 else 'A'
                            else:
                                first_empty_col_letter = 'A'
                            
                            first_empty_cell = f"A{first_empty_row}"  # Always use column A for the first empty cell
                            table_range = f"A1:{chr(65 + last_col_with_data)}{last_row_with_data + 1}"
                            
                            # Find empty cells within the data table area for each column
                            empty_cells_by_column = {}
                            data_start_row = 0  # Find where actual data starts
                            data_end_row = last_row_with_data
                            
                            # Find the first row that has data (skip headers)
                            for row_idx in range(len(df)):
                                if not df.iloc[row_idx].isna().all():
                                    # Check if this row has meaningful data (not just headers)
                                    row_data = df.iloc[row_idx].dropna()
                                    if len(row_data) > 0:
                                        data_start_row = row_idx
                                        break
                            
                            # Analyze each column for empty cells within the data range
                            for col_idx in range(len(df.columns)):
                                if col_idx <= last_col_with_data:  # Only check columns that have data
                                    col_letter = chr(65 + col_idx)  # Convert to A, B, C, D, etc.
                                    empty_cells = []
                                    
                                    # Check for empty cells in this column within the data range
                                    for row_idx in range(data_start_row, data_end_row + 1):
                                        if row_idx < len(df):
                                            cell_value = df.iloc[row_idx, col_idx]
                                            if pd.isna(cell_value) or str(cell_value).strip() == '':
                                                empty_cells.append(f"{col_letter}{row_idx + 1}")  # +1 for 1-indexed
                                    
                                    if empty_cells:
                                        empty_cells_by_column[col_letter] = {
                                            'empty_cells': empty_cells,
                                            'first_empty': empty_cells[0] if empty_cells else None,
                                            'last_empty': empty_cells[-1] if empty_cells else None,
                                            'total_empty': len(empty_cells)
                                        }
                            
                            # Get some sample data from the table
                            sample_data = []
                            for i in range(min(5, len(df))):  # Get first 5 rows as sample
                                row_data = []
                                for j in range(min(5, len(df.columns))):  # Get first 5 columns
                                    cell_value = df.iloc[i, j]
                                    if pd.isna(cell_value):
                                        row_data.append("")
                                    else:
                                        row_data.append(str(cell_value))
                                sample_data.append(row_data)
                        else:
                            last_row_with_data = 0
                            first_empty_row = 1
                            first_empty_cell = "A1"
                            table_range = "A1:A1"
                            sample_data = []
                            empty_cells_by_column = {}
                            data_start_row = 0
                            data_end_row = 0
                        
                        analysis_results[sheet_name] = {
                            'last_row_with_data': last_row_with_data + 1,  # Convert to 1-indexed
                            'first_empty_row': first_empty_row,
                            'first_empty_cell': first_empty_cell,
                            'table_range': table_range,
                            'data_start_row': data_start_row + 1,  # Convert to 1-indexed
                            'data_end_row': data_end_row + 1,  # Convert to 1-indexed
                            'total_rows_with_data': len(non_empty_rows) if 'non_empty_rows' in locals() else 0,
                            'total_columns_with_data': len(non_empty_cols.columns) if 'non_empty_cols' in locals() else 0,
                            'empty_cells_by_column': empty_cells_by_column,
                            'sample_data': sample_data
                        }
                        
                    except Exception as sheet_error:
                        analysis_results[sheet_name] = {
                            'error': f'Could not analyze sheet: {str(sheet_error)}'
                        }
                
                return Response({
                    'status': 'success',
                    'file_label': file_label,
                    'filename': os.path.basename(file_obj.file.name),
                    'sheet_count': len(sheet_names),
                    'sheets': analysis_results
                }, status=status.HTTP_200_OK)
                
            except Exception as e:
                return Response({
                    'message': f'Error analyzing file: {str(e)}',
                    'status': 'error'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Comparison.DoesNotExist:
            return Response({
                'message': 'Comparison not found or you do not have permission to access it',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'message': f'Error: {str(e)}',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


