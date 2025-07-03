import json
import os
import io
import re
import pandas as pd
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from ..models import FileUpload, Comparison
from ..serializers import ComparisonSerializer
from ..permissions import IsAdmin
from ..funtions import compare_excel_sheets, preprocess_results
from ..encryption import decrypt_file
from ..funtions import open_file












def find_formatted_table_boundaries(worksheet, start_cell="A1"):
    """
    Find the boundaries of a formatted table (not Excel Table object) by detecting 
    where data starts and ends based on cell content patterns.
    """
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
    
    # Parse start cell
    col_letter, row_num = coordinate_from_string(start_cell)
    start_col = column_index_from_string(col_letter)
    start_row = row_num
    
    # First, find where the actual table data starts by looking for headers
    table_start_row = start_row
    table_start_col = start_col
    
    # Look for the first row that has multiple non-empty cells (likely headers)
    for row in range(start_row, min(start_row + 20, worksheet.max_row + 1)):
        non_empty_count = 0
        first_col_with_data = None
        
        # Check first 10 columns for data
        for col in range(start_col, min(start_col + 10, worksheet.max_column + 1)):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                non_empty_count += 1
                if first_col_with_data is None:
                    first_col_with_data = col
        
        # If we find a row with multiple data cells, this is likely our table start
        if non_empty_count >= 2:
            table_start_row = row
            table_start_col = first_col_with_data
            break
    
    # Find the last row with data
    last_row = table_start_row
    consecutive_empty_rows = 0
    
    for row in range(table_start_row + 1, worksheet.max_row + 1):
        row_has_data = False
        
        # Check if any cell in this row has data
        for col in range(table_start_col, min(table_start_col + 10, worksheet.max_column + 1)):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                row_has_data = True
                break
        
        if row_has_data:
            last_row = row
            consecutive_empty_rows = 0
        else:
            consecutive_empty_rows += 1
            # If we hit 3 consecutive empty rows, we've likely reached the end
            if consecutive_empty_rows >= 3:
                break
    
    # Find the last column with data
    last_col = table_start_col
    consecutive_empty_cols = 0
    
    for col in range(table_start_col, worksheet.max_column + 1):
        col_has_data = False
        
        # Check if any cell in this column has data
        for row in range(table_start_row, min(last_row + 1, table_start_row + 100)):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                col_has_data = True
                break
        
        if col_has_data:
            last_col = col
            consecutive_empty_cols = 0
        else:
            consecutive_empty_cols += 1
            # If we hit 2 consecutive empty columns, we've likely reached the end
            if consecutive_empty_cols >= 2:
                break
    
    # Convert back to Excel notation
    start_cell_final = f"{get_column_letter(table_start_col)}{table_start_row}"
    end_cell = f"{get_column_letter(last_col)}{last_row}"
    
    return f"{start_cell_final}:{end_cell}", table_start_row, last_row, table_start_col, last_col





class AddComparisonView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, file1_id, file2_id):
        sheet1 = request.data.get("sheet1")
        sheet2 = request.data.get("sheet2")
        range1 = request.data.get("range1", "A7:F99")
        range2 = request.data.get("range2", "A7:F99")
        column1 = request.data.get("column1")
        column2 = request.data.get("column2")

        if not file1_id or not file2_id:
            return Response(
                {"error": "file1, file2, and results are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if column1 != column2:
            return Response(
                {"error": "primary key in file 1 not equal file2 primary key"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            file1 = FileUpload.objects.get(id=file1_id, user=request.user)
            file2 = FileUpload.objects.get(id=file2_id, user=request.user)

            if file1.id == file2.id:
                return Response(
                    {"error": "file1 and file2 must be different files."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            primary_columns = None
            if column1 and column2:
                primary_columns = [column1, column2]

            results_json = compare_excel_sheets(
                file1.file.path,
                file2.file.path,
                sheet_name=sheet1,
                range_file1=range1,
                range_file2=range2,
                PrimaryColumn=primary_columns
            )

            results_json = preprocess_results(results_json)

            comparison = Comparison.objects.create(
                file1=file1,
                file2=file2,
                results=json.dumps(results_json),
                user_id=request.user.id,
                range1=range1,
                range2=range2, 
                column1=column1,
                column2=column2,
            )

            serializer = ComparisonSerializer(comparison)
            return Response(
                {"message": "Comparison created successfully", "data": serializer.data},
                status=status.HTTP_201_CREATED,
            )
        except FileUpload.DoesNotExist:
            return Response(
                {"error": "One or both files not found or do not belong to the user."},
                status=status.HTTP_404_NOT_FOUND,
            )


class DeleteComparisonView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, comparison_id):
        comparison = Comparison.objects.filter(id=comparison_id).first()
        if not comparison:
            return Response({"message": "Comparison not found."}, status=status.HTTP_404_NOT_FOUND)

        comparison.delete()
        return Response({"message": "Comparison deleted successfully."}, status=status.HTTP_200_OK)


class ListUserComparisonView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        comparisons = Comparison.objects.filter(user=request.user)
    
        search_query = request.GET.get('search', '').strip()
        if search_query:
            comparisons = comparisons.filter(
                Q(file1__file__icontains=search_query) |
                Q(file2__file__icontains=search_query)
            )

        comparison_id = request.GET.get('comparison_id', '').strip()
        if comparison_id:
            comparisons = comparisons.filter(id=comparison_id)

        # Date filtering
        date_filter_type = request.GET.get('date_filter_type', '')
        if date_filter_type == 'range':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            if start_date and end_date:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                end_date = end_date_obj.strftime('%Y-%m-%d')
                comparisons = comparisons.filter(timestamp__range=[start_date, end_date])
        elif date_filter_type == 'day' and request.GET.get('filter_date'):
            filter_date = request.GET.get('filter_date')
            comparisons = comparisons.filter(timestamp__date=filter_date)
        elif date_filter_type == 'month' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET.get('filter_date'), '%Y-%m-%d')
            comparisons = comparisons.filter(
                timestamp__year=filter_date.year,
                timestamp__month=filter_date.month
            )
        elif date_filter_type == 'year' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET.get('filter_date'), '%Y-%m-%d')
            comparisons = comparisons.filter(timestamp__year=filter_date.year)

        paginator = PageNumberPagination()
        paginator.page_size = 10
        results_page = paginator.paginate_queryset(comparisons, request)
        serializer = ComparisonSerializer(results_page, many=True)
        data = serializer.data
        for i, comp in enumerate(results_page):
            data[i]['file1_name'] = comp.file1.file.name
            data[i]['file2_name'] = comp.file2.file.name
            data[i]['user'] = comp.user.username
        return Response({
            'results': data,
            'current_page': paginator.page.number,
            'total_pages': paginator.page.paginator.num_pages
        }, status=status.HTTP_200_OK)


class ListAllComparisonView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        comparisons = Comparison.objects.all().select_related('file1', 'file2', 'user')

        search_query = request.GET.get('search', '').strip()
        if search_query:
            comparisons = comparisons.filter(
                Q(file1__file__icontains=search_query) |
                Q(file2__file__icontains=search_query)
            )

        comparison_id = request.GET.get('comparison_id', '').strip()
        if comparison_id:
            comparisons = comparisons.filter(id=comparison_id)

        # Date filtering
        date_filter_type = request.GET.get('date_filter_type', '')
        if date_filter_type == 'range':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            if start_date and end_date:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                comparisons = comparisons.filter(timestamp__range=[start_date, end_date_obj.strftime('%Y-%m-%d')])
        elif date_filter_type == 'day' and request.GET.get('filter_date'):
            comparisons = comparisons.filter(timestamp__date=request.GET['filter_date'])
        elif date_filter_type == 'month' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET['filter_date'], '%Y-%m-%d')
            comparisons = comparisons.filter(timestamp__year=filter_date.year, timestamp__month=filter_date.month)
        elif date_filter_type == 'year' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET['filter_date'], '%Y-%m-%d')
            comparisons = comparisons.filter(timestamp__year=filter_date.year)

        comparisons = comparisons.order_by('user__id', '-timestamp')

        paginator = PageNumberPagination()
        paginator.page_size = 10
        results_page = paginator.paginate_queryset(comparisons, request)
        data = []
        for comp in results_page:
            data.append({
                'id': comp.id,
                'file1_name': comp.file1.file.name,
                'file2_name': comp.file2.file.name,
                'user': comp.user.username,
                'timestamp': comp.timestamp
            })
        return paginator.get_paginated_response(data)


class ComparisonDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, comparison_id):
        if request.user.role == 'admin':
            comparison = Comparison.objects.filter(id=comparison_id).first()
        else:
            comparison = Comparison.objects.filter(id=comparison_id, user=request.user).first()

        if not comparison:
            return Response({"error": "Comparison not found or you don't have permission to view it"}, 
                           status=status.HTTP_404_NOT_FOUND)

        data = {
            'id': comparison.id,
            'file1_name': comparison.file1.file.name,
            'file2_name': comparison.file2.file.name,
            'user': comparison.user.username,
            'timestamp': comparison.timestamp,
            'results': comparison.results
        }
        return Response(data, status=status.HTTP_200_OK)


def column_letter_to_index(col_letter):
    """Convert column letter (A, B, C, etc.) to 0-based index"""
    result = 0
    for char in col_letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1
def get_excel_column_name(column_index):
    """Convert 0-based column index to Excel column letter(s)"""
    result = ""
    while column_index >= 0:
        result = chr(column_index % 26 + ord('A')) + result
        column_index = column_index // 26 - 1
        if column_index < 0:
            break
    return result
def get_next_column_letter(col_letter):
    """Get the next column letter (A->B, Z->AA, etc.)"""
    col_letter = col_letter.upper()
    
    # Convert to list for easier manipulation
    letters = list(col_letter)
    
    # Start from the rightmost character
    i = len(letters) - 1
    
    while i >= 0:
        if letters[i] != 'Z':
            # Simple case: increment the letter
            letters[i] = chr(ord(letters[i]) + 1)
            break
        else:
            # This letter is Z, so it becomes A and we carry over
            letters[i] = 'A'
            if i == 0:
                # We've carried over past the first letter, so add a new A at the beginning
                letters.insert(0, 'A')
                break
            # Continue to the next letter to the left
            i -= 1
    
    return ''.join(letters)
def safe_assign_to_dataframe(df, row_index, column_index, value, column_name=None):
    """
    Safely assign a value to df.iloc[row_index, column_index].
    Expands the DataFrame if necessary.
    """
    # Ensure we have enough rows
    if row_index >= len(df):
        # Add missing rows
        missing_rows = row_index - len(df) + 1
        for _ in range(missing_rows):
            df.loc[len(df)] = [None] * len(df.columns)
    
    # Ensure we have enough columns
    current_columns = len(df.columns)
    if column_index >= current_columns:
        # Add missing columns
        for i in range(current_columns, column_index + 1):
            if column_name and i == column_index:
                col_name = column_name
            else:
                col_name = get_excel_column_name(i)
            df[col_name] = None
    
    # Now safely assign the value
    df.iloc[row_index, column_index] = value
    return df




class DownloadComparisonExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, comparison_id, file_number=1):
        """
        Download the Excel file from a comparison.
        file_number: 1 for first file, 2 for second file
        """
        try:
            # For admin, allow access to any comparison
            if request.user.role == 'admin':
                comparison = Comparison.objects.get(id=comparison_id)
            else:
                # For regular users, only allow access to their own comparisons
                comparison = Comparison.objects.get(id=comparison_id, user=request.user)
                
            # Determine which file to download
            if file_number == 1:
                file_obj = comparison.file1
                file_label = "source"
            elif file_number == 2:
                file_obj = comparison.file2
                file_label = "target"
            else:
                return Response({
                    'message': 'Invalid file number. Use 1 for source file or 2 for target file.',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get the file path and comparison data
            range1 = comparison.range1
            range2 = comparison.range2
            result = comparison.results
            
            if isinstance(result, str):
                result_data = json.loads(result)
            else:
                result_data = result
                
            print(f"Range1: {range1}, Range2: {range2}")
            print(f"Comparison Results: {result}")
            
            file_path = file_obj.file.path
            
            # Check if file exists
            if not os.path.exists(file_path):
                return Response({
                    'message': f'File not found: {file_label} file for comparison {comparison_id}',
                    'status': 'error'
                }, status=status.HTTP_404_NOT_FOUND)
                
            # Decrypt the file temporarily for download
            try:
                # Check the file size before decryption
                file_size = os.path.getsize(file_path)
                print(f"Original file size: {file_size} bytes")
                
                # First, try to read the file directly to see if it's already decrypted
                with open(file_path, 'rb') as f:
                    first_bytes = f.read(10)
                    print(f"First 10 bytes of file: {first_bytes}")
                
                # Check if the file is already in Excel format (not encrypted)
                if first_bytes.startswith(b'PK') or first_bytes.startswith(b'\xd0\xcf\x11\xe0'):
                    print("File appears to be unencrypted Excel format")
                    # Read the entire file as-is
                    with open(file_path, 'rb') as f:
                        decrypted_content = f.read()
                    print(f"Direct file read size: {len(decrypted_content)} bytes")
                else:
                    print("File appears to be encrypted, attempting decryption...")
                    # Try to decrypt the file
                    try:
                        decrypted_result = decrypt_file(file_path)
                        print(f"Decrypt function returned: type={type(decrypted_result)}, length={len(str(decrypted_result))}")
                        
                        # Check if decrypt_file returned a path (string) instead of content
                        if isinstance(decrypted_result, str):
                            if os.path.exists(decrypted_result):
                                print(f"Decrypt function returned a file path: {decrypted_result}")
                                print("Reading the decrypted file content...")
                                with open(decrypted_result, 'rb') as f:
                                    decrypted_content = f.read()
                                print(f"File content read successfully: {len(decrypted_content)} bytes")
                            else:
                                print("Decrypt function returned a string, but it's not a valid path")
                                print(f"First 100 characters: {decrypted_result[:100]}")
                                # Try to encode the string as bytes
                                decrypted_content = decrypted_result.encode('latin-1')
                        else:
                            # Decrypt function returned bytes directly
                            decrypted_content = decrypted_result
                            
                        print(f"Final decrypted content size: {len(decrypted_content)} bytes")
                        print(f"Final decrypted content type: {type(decrypted_content)}")
                        
                    except Exception as decrypt_error:
                        print(f"Decryption failed: {str(decrypt_error)}")
                        return Response({
                            'message': f'Error decrypting file: {str(decrypt_error)}',
                            'status': 'error'
                        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
                # Ensure decrypted_content is bytes
                if isinstance(decrypted_content, str):
                    print("Converting string to bytes using latin-1 encoding")
                    try:
                        decrypted_content = decrypted_content.encode('latin-1')
                    except UnicodeEncodeError:
                        print("Failed to encode with latin-1, trying utf-8")
                        try:
                            decrypted_content = decrypted_content.encode('utf-8')
                        except UnicodeEncodeError:
                            print("Both encoding methods failed, file might be corrupted")
                            return Response({
                                'message': 'File encoding error - file may be corrupted',
                                'status': 'error'
                            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
                # Verify we have bytes now
                if not isinstance(decrypted_content, bytes):
                    return Response({
                        'message': f'Error: Decrypted content is {type(decrypted_content)}, expected bytes',
                        'status': 'error'
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                    
                # Verify the decrypted content starts with valid Excel magic bytes
                if decrypted_content.startswith(b'PK'):
                    print("File appears to be a valid zip-based format (Excel .xlsx)")
                elif decrypted_content.startswith(b'\xd0\xcf\x11\xe0'):
                    print("File appears to be OLE format (Excel .xls)")
                else:
                    print(f"Warning: File doesn't start with expected Excel headers. First 10 bytes: {decrypted_content[:10]}")
                    print("File might still be valid - proceeding with download")
                    
                # Modify the existing Excel file structure
                try:
                    # Create a BytesIO object from the decrypted content
                    excel_buffer = io.BytesIO(decrypted_content)
                    
                    # Load the workbook using openpyxl to preserve original structure
                    workbook = load_workbook(excel_buffer)
                    
                    # Get the first worksheet (or specify sheet name if needed)
                    worksheet = workbook.active  # or workbook['SheetName']
                    
                    # Parse the range
                    rangee = range2.split(':')
                    Range1 = rangee[0].strip().upper()
                    Range2 = rangee[1].strip().upper() if len(rangee) > 1 else None

                    match = re.match(r'([A-Z]+)(\d+)', Range1)
                    match2 = re.match(r'([A-Z]+)(\d+)', Range2)
                    
                    if match and match2:
                        col_letter = match.group(1)
                        row_number = int(match.group(2))
                        col_letter2 = match2.group(1)
                        row_number2 = int(match2.group(2))
                        next_column = get_next_column_letter(col_letter2)
                        
                        print(f"Current column: {col_letter2}, Next column: {next_column}")

                        col_letter_index = column_letter_to_index(col_letter)
                        next_column_index = column_letter_to_index(next_column)
                        
                        # Get comparison results
                        added_row = result_data.get("rows_added", [])
                        removed_row = result_data.get("rows_removed", [])
                        changed_values = result_data.get("value_diff", [])
                        column= comparison.column1
                        
                        # Create a set of rows to mark as unmatched for efficiency
                        unmatched_rows = set()
                        workbook2 = load_workbook(excel_buffer)
                        worksheet2 = workbook2.active
                        store_matched_rows = []

                        for row_num in range(row_number+2, row_number2): # +1 to skip header row
                            cell_value = worksheet2.cell(row=row_num, column=col_letter_index+1).value
                            if cell_value is None or cell_value == '':
                                continue
                            store_matched_rows.append(row_num)
                            print(cell_value) 
                        
                        # Process added rows
                        for item in added_row:                            
                              if column in item and column in item and item[column] is not None:
                                unmatched_rows.add(item['_OriginalRow'])
                        

                        # Process changed values
                        for item in changed_values:
                            if file_number == 1 and 'excel_row_file1' in item:
                                unmatched_rows.add(item['excel_row_file1'])
                            elif file_number == 2 and 'excel_row_file2' in item:
                                unmatched_rows.add(item['excel_row_file2'])
                        
                        # Add header for the new column if it doesn't exist
                        header_cell = worksheet.cell(row=row_number, column=next_column_index + 1)
                        if header_cell.value is None or header_cell.value == '':
                            header_cell.value = "Comparison Status"
                            header_cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Yellow

                        

                        total_matched = 0
                        total_unmatched = len(unmatched_rows)   
                        # Mark unmatched rows in the new column
                        for row_target in range(row_number + 2, row_number2 ):  # +1 to skip header row
                            cell = worksheet.cell(row=row_target, column=next_column_index + 1)
                            
                            if row_target in unmatched_rows:
                                cell.value = "unmatched"
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
                                print(f"Row {row_target} in column {next_column} marked as unmatched")

                            elif row_target in store_matched_rows:
                                cell.value = "match"
                                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                                total_matched += 1
                                print(f"Row {row_target} in column {next_column} marked as match")

                        total_rows_with_data = len(store_matched_rows)
                        matched_percentage = (total_matched / total_rows_with_data * 100) if total_rows_with_data > 0 else 0
                        unmatched_percentage = (total_unmatched / total_rows_with_data * 100) if total_rows_with_data > 0 else 0

                        cell=worksheet.cell(row=row_number , column=next_column_index + 2)
                        cell.value = "Total matched"
                        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Medium gray for header



                        cell = worksheet.cell(row=row_number + 3, column=next_column_index + 2)
                        cell.value = f"{total_matched} ({matched_percentage:.1f}%)"
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                          # Light gray

                        cell = worksheet.cell(row=row_number, column=next_column_index + 3)
                        cell.value = "Total unmatched"
                        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

                        cell = worksheet.cell(row=row_number + 3, column=next_column_index + 3)
                        cell.value = f"{total_unmatched} ({unmatched_percentage:.1f}%)"
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow

                        comparison_status_col = get_excel_column_name(next_column_index)
                        worksheet.column_dimensions[comparison_status_col].width = 20

                        comparison_status_col = get_excel_column_name(next_column_index+1)
                        worksheet.column_dimensions[comparison_status_col].width = 20

                        comparison_status_col = get_excel_column_name(next_column_index+2)
                        worksheet.column_dimensions[comparison_status_col].width = 20
                            
                    
                    # Save the modified workbook to a new BytesIO buffer
                    modified_excel_buffer = io.BytesIO()
                    workbook.save(modified_excel_buffer)
                    
                    # Get the modified content
                    modified_excel_buffer.seek(0)
                    final_content = modified_excel_buffer.getvalue()
                    
                    print(f"Modified Excel content size: {len(final_content)} bytes")
                    
                except Exception as excel_modify_error:
                    print(f"Error modifying Excel file: {str(excel_modify_error)}")
                    # Fallback to original content if modification fails
                    final_content = decrypted_content
                
                # Get the original filename without path
                original_filename = os.path.basename(file_obj.file.name)
                
                # Ensure proper Excel content type
                if original_filename.lower().endswith('.xlsx'):
                    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                elif original_filename.lower().endswith('.xls'):
                    content_type = 'application/vnd.ms-excel'
                elif original_filename.lower().endswith('.csv'):
                    content_type = 'text/csv'
                else:
                    # Default to Excel format
                    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    
                # Create the response with the modified content
                response = HttpResponse(final_content, content_type=content_type)
                response['Content-Disposition'] = f'attachment; filename="{original_filename}"'
                response['Content-Length'] = len(final_content)
                response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                response['Pragma'] = 'no-cache'
                response['Expires'] = '0'
                
                print(f"Sending response with {len(final_content)} bytes, content-type: {content_type}")
                return response
                
            except Exception as e:
                return Response({
                    'message': f'Error processing file: {str(e)}',
                    'status': 'error'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Comparison.DoesNotExist:
            return Response({
                'message': 'Comparison not found or you do not have permission to access it',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'message': f'Error: {str(e)}',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AnalyzeExcelTableView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, comparison_id, file_number=1):
        """Analyze Excel table structure and return the first empty cell after the table."""
        try:
            if request.user.role == 'admin':
                comparison = Comparison.objects.get(id=comparison_id)
            else:
                comparison = Comparison.objects.get(id=comparison_id, user=request.user)
            
            if file_number == 1:
                file_obj = comparison.file1
                file_label = "source"
            elif file_number == 2:
                file_obj = comparison.file2
                file_label = "target"
            else:
                return Response({
                    'message': 'Invalid file number. Use 1 for source file or 2 for target file.',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            file_path = file_obj.file.path
            
            if not os.path.exists(file_path):
                return Response({
                    'message': f'File not found: {file_label} file for comparison {comparison_id}',
                    'status': 'error'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # File analysis logic would go here
            return Response({
                'status': 'success',
                'file_label': file_label,
                'filename': os.path.basename(file_obj.file.name),
                'sheet_count': 0,
                'sheets': {}
            }, status=status.HTTP_200_OK)
                
        except Comparison.DoesNotExist:
            return Response({
                'message': 'Comparison not found or you do not have permission to access it',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'message': f'Error: {str(e)}',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)














































class MakeCompareView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, file1_id, file2_id):
        sheet1 = request.data.get("sheet1")
        sheet2 = request.data.get("sheet2")
        range1 = request.data.get("range1", "A7:F99")
        range2 = request.data.get("range2", "A7:F99")
        primary_columns1 = request.data.get("column1")
        primary_columns2 = request.data.get("column2")

        if not file1_id or not file2_id:
            return Response(
                {"error": "file1, file2, and results are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        if primary_columns1 != primary_columns2:
            return Response(
                {"error": "primary key in file 1 not equal file2 primary key"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        try:
            file1 = FileUpload.objects.get(id=file1_id, user=request.user)
            file2 = FileUpload.objects.get(id=file2_id, user=request.user)

            if file1.id == file2.id:
                return Response(
                    {"error": "file1 and file2 must be different files."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            primary_columns = None
            if primary_columns1 and primary_columns2:
                primary_columns = [primary_columns1, primary_columns2]
           
            # Decrypt files
            decrypted_file1 = decrypt_file(file1.file.path)
            decrypted_file2 = decrypt_file(file2.file.path)
        except Exception as e:
            return Response(
                {"error": f"Error decrypting files: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

            # Use load_workbook instead of open_file
        try:
                # Load workbooks with openpyxl
                if isinstance(decrypted_file1, str) and os.path.exists(decrypted_file1):
                    workbook1 = load_workbook(decrypted_file1, data_only=True)
                else:
                    workbook1 = load_workbook(io.BytesIO(decrypted_file1), data_only=True)
                
                if isinstance(decrypted_file2, str) and os.path.exists(decrypted_file2):
                    workbook2 = load_workbook(decrypted_file2, data_only=True)
                else:
                    workbook2 = load_workbook(io.BytesIO(decrypted_file2), data_only=True)

                # Get worksheets
                if sheet1:
                    worksheet1 = workbook1[sheet1]
                else:
                    worksheet1 = workbook1.active

                if sheet2:
                    worksheet2 = workbook2[sheet2]
                else:
                    worksheet2 = workbook2.active

                print(f"Range1: {range1}")
                print(f"Range2: {range2}")

                # Extract data from range1 (e.g., "A6:C31")
                print("\n=== ITERATING THROUGH RANGE1 ===")
                print(f"File1 - Range: {range1}")
                
                range1_data = []
                row_count = 0
                for row in worksheet1[range1]:
                    row_values = [cell.value for cell in row]
                    range1_data.append(row_values)
                    
                    # Print each row
                    print(f"Row {row_count + 1}: {row_values}")
                    
                    # Print each cell in the row with column info
                    for col_index, cell_value in enumerate(row_values):
                        col_letter = chr(ord('A') + col_index)  # Convert to A, B, C, etc.
                        print(f"  Column {col_letter}: {cell_value}")
                    
                    print("-" * 40)
                    row_count += 1

                # Extract data from range2 (e.g., "A6:C31")
                print("\n=== ITERATING THROUGH RANGE2 ===")
                print(f"File2 - Range: {range2}")
                
                range2_data = []
                row_count = 0
                for row in worksheet2[range2]:
                    row_values = [cell.value for cell in row]
                    range2_data.append(row_values)
                    
                    # Print each row
                    print(f"Row {row_count + 1}: {row_values}")
                    
                    # Print each cell in the row with column info
                    for col_index, cell_value in enumerate(row_values):
                        col_letter = chr(ord('A') + col_index)  # Convert to A, B, C, etc.
                        print(f"  Column {col_letter}: {cell_value}")
                    
                    print("-" * 40)
                    row_count += 1
                
                # Convert to DataFrames
                df1 = pd.DataFrame(range1_data)
                df2 = pd.DataFrame(range2_data)

                print(f"DataFrame1 shape: {df1.shape}")
                print(f"DataFrame2 shape: {df2.shape}")
                
                # Display first few rows of raw data for debugging
                print("First 5 rows of df1 (raw data):")
                print(df1.head())
                
                # Set headers and process data
                if not df1.empty and len(df1) > 0:
                    # Use first row as headers
                    df1.columns = df1.iloc[0]
                    df1 = df1.drop(df1.index[0])
                    df1.reset_index(drop=True, inplace=True)
                    
                    # Clean column names (remove None values)
                    df1.columns = [str(col) if col is not None else f"Column_{i}" for i, col in enumerate(df1.columns)]
                    
                    print("DataFrame 1 Headers:", df1.columns.tolist())
                    print(f"DataFrame 1 data rows: {len(df1)}")
                    
                    # Print first few rows with data
                    print("First 3 data rows from df1:")
                    for index, row in df1.head(3).iterrows():
                        print(f"Row {index}:")
                        for column in df1.columns:
                            if pd.notna(row[column]) and str(row[column]).strip():
                                print(f"  {column}: {row[column]}")
                        print("-" * 40)

                if not df2.empty and len(df2) > 0:
                    # Use first row as headers
                    df2.columns = df2.iloc[0]
                    df2 = df2.drop(df2.index[0])
                    df2.reset_index(drop=True, inplace=True)
                    
                    # Clean column names
                    df2.columns = [str(col) if col is not None else f"Column_{i}" for i, col in enumerate(df2.columns)]
                    
                    print("DataFrame 2 Headers:", df2.columns.tolist())
                    print(f"DataFrame 2 data rows: {len(df2)}")

        except Exception as workbook_error:
                return Response(
                    {"error": f"Error loading Excel files: {str(workbook_error)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
































