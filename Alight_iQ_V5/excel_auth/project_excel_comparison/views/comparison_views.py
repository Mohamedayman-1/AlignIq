import json
import os
import io
import re
import pandas as pd
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from ..models import FileUpload, Comparison
from ..serializers import ComparisonSerializer
from ..permissions import IsAdmin
from ..funtions import compare_excel_sheets, preprocess_results
from ..encryption import decrypt_file
from ..funtions import open_file












def find_formatted_table_boundaries(worksheet, start_cell="A1"):
    """
    Find the boundaries of a formatted table (not Excel Table object) by detecting 
    where data starts and ends based on cell content patterns.
    """
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
    
    # Parse start cell
    col_letter, row_num = coordinate_from_string(start_cell)
    start_col = column_index_from_string(col_letter)
    start_row = row_num
    
    # First, find where the actual table data starts by looking for headers
    table_start_row = start_row
    table_start_col = start_col
    
    # Look for the first row that has multiple non-empty cells (likely headers)
    for row in range(start_row, min(start_row + 20, worksheet.max_row + 1)):
        non_empty_count = 0
        first_col_with_data = None
        
        # Check first 10 columns for data
        for col in range(start_col, min(start_col + 10, worksheet.max_column + 1)):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                non_empty_count += 1
                if first_col_with_data is None:
                    first_col_with_data = col
        
        # If we find a row with multiple data cells, this is likely our table start
        if non_empty_count >= 2:
            table_start_row = row
            table_start_col = first_col_with_data
            break
    
    # Find the last row with data
    last_row = table_start_row
    consecutive_empty_rows = 0
    
    for row in range(table_start_row + 1, worksheet.max_row + 1):
        row_has_data = False
        
        # Check if any cell in this row has data
        for col in range(table_start_col, min(table_start_col + 10, worksheet.max_column + 1)):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                row_has_data = True
                break
        
        if row_has_data:
            last_row = row
            consecutive_empty_rows = 0
        else:
            consecutive_empty_rows += 1
            # If we hit 3 consecutive empty rows, we've likely reached the end
            if consecutive_empty_rows >= 3:
                break
    
    # Find the last column with data
    last_col = table_start_col
    consecutive_empty_cols = 0
    
    for col in range(table_start_col, worksheet.max_column + 1):
        col_has_data = False
        
        # Check if any cell in this column has data
        for row in range(table_start_row, min(last_row + 1, table_start_row + 100)):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                col_has_data = True
                break
        
        if col_has_data:
            last_col = col
            consecutive_empty_cols = 0
        else:
            consecutive_empty_cols += 1
            # If we hit 2 consecutive empty columns, we've likely reached the end
            if consecutive_empty_cols >= 2:
                break
    
    # Convert back to Excel notation
    start_cell_final = f"{get_column_letter(table_start_col)}{table_start_row}"
    end_cell = f"{get_column_letter(last_col)}{last_row}"
    
    return f"{start_cell_final}:{end_cell}", table_start_row, last_row, table_start_col, last_col







































class AddComparisonView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, file1_id, file2_id):
        sheet1 = request.data.get("sheet1")
        sheet2 = request.data.get("sheet2")
        range1 = request.data.get("range1", "A7:F99")
        range2 = request.data.get("range2", "A7:F99")
        column1 = request.data.get("column1")
        column2 = request.data.get("column2")

        if not file1_id or not file2_id:
            return Response(
                {"error": "file1, file2, and results are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if column1 != column2:
            return Response(
                {"error": "primary key in file 1 not equal file2 primary key"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            file1 = FileUpload.objects.get(id=file1_id, user=request.user)
            file2 = FileUpload.objects.get(id=file2_id, user=request.user)

            if file1.id == file2.id:
                return Response(
                    {"error": "file1 and file2 must be different files."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            primary_columns = None
            if column1 and column2:
                primary_columns = [column1, column2]

            # Decrypt and load files
            try:
                decrypted_file1 = decrypt_file(file1.file.path)
                decrypted_file2 = decrypt_file(file2.file.path)

                # Load workbooks
                if isinstance(decrypted_file1, str) and os.path.exists(decrypted_file1):
                    workbook1 = load_workbook(decrypted_file1, data_only=True)
                else:
                    workbook1 = load_workbook(io.BytesIO(decrypted_file1), data_only=True)
                
                if isinstance(decrypted_file2, str) and os.path.exists(decrypted_file2):
                    workbook2 = load_workbook(decrypted_file2, data_only=True)
                else:
                    workbook2 = load_workbook(io.BytesIO(decrypted_file2), data_only=True)

                # Get worksheets
                worksheet1 = workbook1[sheet1] if sheet1 else workbook1.active
                worksheet2 = workbook2[sheet2] if sheet2 else workbook2.active

                # Extract formatted data from ranges
                file1_formatted_data = self.extract_formatted_data(worksheet1, range1, "file1")
                file2_formatted_data = self.extract_formatted_data(worksheet2, range2, "file2")

                # Run original comparison
                results_json = compare_excel_sheets(
                    file1.file.path,
                    file2.file.path,
                    sheet_name=sheet1,
                    range_file1=range1,
                    range_file2=range2,
                    PrimaryColumn=primary_columns
                )

                results_json = preprocess_results(results_json)

                # Create structured comparison data
                structured_results = self.create_structured_results(
                    file1_formatted_data, 
                    file2_formatted_data, 
                    results_json, 
                    column1
                )

                comparison = Comparison.objects.create(
                    file1=file1,
                    file2=file2,
                    results=json.dumps(structured_results),
                    user_id=request.user.id,
                    range1=range1,
                    range2=range2, 
                    column1=column1,
                    column2=column2,
                )

                # Return structured data for frontend
                return Response({
                    "message": "Comparison created successfully",
                    "comparison_id": comparison.id,
                    "data": structured_results
                }, status=status.HTTP_201_CREATED)

            except Exception as processing_error:
                return Response(
                    {"error": f"Error processing files: {str(processing_error)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        except FileUpload.DoesNotExist:
            return Response(
                {"error": "One or both files not found or do not belong to the user."},
                status=status.HTTP_404_NOT_FOUND,
            )

    def extract_formatted_data(self, worksheet, range_str, file_source):
        """Extract data from worksheet range and format it for frontend"""
        data = []
        headers = []
        
        # Convert range to actual rows for processing
        all_rows = list(worksheet[range_str])
        
        if len(all_rows) == 0:
            return {
                "headers": [],
                "rows": [],
                "total_rows": 0,
                "range": range_str,
                "source": file_source,
                "header_analysis": {
                    "detected_pattern": "no_data",
                    "confidence": "low",
                    "description": "No data found in range"
                }
            }
        
        # Analyze header structure with error handling
        try:
            header_structure = self.analyze_header_structure(all_rows)
        except Exception as e:
            print(f"Error analyzing header structure: {e}")
            header_structure = {
                "detected_multi_row": False,
                "confidence": 0.5,
                "analysis_details": "Error during analysis"
            }
        
        # Detect and combine multi-row headers with error handling
        try:
            headers = self.detect_and_combine_headers(all_rows)
        except Exception as e:
            print(f"Error detecting headers: {e}")
            # Fallback to simple column naming
            headers = [f"Column_{i+1}" for i in range(len(all_rows[0]) if all_rows else 0)]
        
        # Validate header detection results with error handling
        try:
            validation_results = self.validate_header_detection(headers, header_structure["analysis_details"])
        except Exception as e:
            print(f"Error validating headers: {e}")
            validation_results = {
                "confidence": 0.5,
                "warnings": ["Error during validation"],
                "recommendations": ["Manual review recommended"]
            }
        
        # Process data rows (skip header rows)
        data_start_index = self.find_data_start_row(all_rows)
        
        for i in range(data_start_index, len(all_rows)):
            row = all_rows[i]
            row_values = [cell.value for cell in row]
            
            # Skip completely empty rows
            if all(val is None or str(val).strip() == '' for val in row_values):
                continue
            
            row_data = {
                "row_number": i - data_start_index + 1,
                "excel_row": row[0].row,  # Actual Excel row number
                "source": file_source,
                "data": {},
                "match_status": "pending",  # Will be updated based on comparison
                "has_changes": False,
                "changed_columns": []
            }
            
            # Create column data
            for j, (header, value) in enumerate(zip(headers, row_values)):
                row_data["data"][header] = {
                    "value": value,
                    "column_index": j,
                    "column_letter": chr(ord('A') + j),
                    "is_changed": False
                }
            
            data.append(row_data)
        
        return {
            "headers": headers,
            "rows": data,
            "total_rows": len(data),
            "range": range_str,
            "source": file_source,
            "header_analysis": {
                "data_start_row": data_start_index + 1,
                "header_rows_used": data_start_index,
                "detected_pattern": header_structure["detected_pattern"],
                "confidence": header_structure["confidence"],
                "description": header_structure["description"],
                "analysis_details": header_structure["analysis_details"],
                "validation": validation_results
            }
        }

    def detect_and_combine_headers(self, all_rows, max_header_rows=3):
        """Detect and combine multi-row headers automatically with comprehensive error handling"""
        try:
            if len(all_rows) == 0:
                return []
            
            # Get the number of columns from the first row
            num_columns = len(all_rows[0])
            
            # Analyze first few rows to detect header pattern
            header_analysis = []
            for row_idx in range(min(max_header_rows, len(all_rows))):
                try:
                    row_values = [cell.value for cell in all_rows[row_idx]]
                    row_analysis = {
                        'row_index': row_idx,
                        'values': row_values,
                        'non_empty_count': sum(1 for val in row_values if val is not None and str(val).strip()),
                        'has_text': any(isinstance(val, str) and not str(val).replace('.', '').replace('-', '').isdigit() 
                                       for val in row_values if val is not None),
                        'has_numbers_only': all(self.is_numeric_or_date(val) for val in row_values if val is not None and str(val).strip()),
                        'has_years': any(self.is_year_value(val) for val in row_values if val is not None),
                        'has_months': any(self.is_month_value(val) for val in row_values if val is not None)
                    }
                    header_analysis.append(row_analysis)
                except Exception as e:
                    print(f"Error analyzing row {row_idx}: {e}")
                    # Fallback analysis for problematic rows
                    row_analysis = {
                        'row_index': row_idx,
                        'values': [],
                        'non_empty_count': 0,
                        'has_text': True,  # Assume it's text to be safe
                        'has_numbers_only': False,
                        'has_years': False,
                        'has_months': False
                    }
                    header_analysis.append(row_analysis)
            
            # Determine how many rows are headers
            header_row_count = self.determine_header_row_count(header_analysis)
            
            # Combine headers from detected header rows
            combined_headers = []
            for col_idx in range(num_columns):
                try:
                    header_parts = []
                    
                    for row_idx in range(header_row_count):
                        if row_idx < len(all_rows):
                            cell_value = all_rows[row_idx][col_idx].value
                            if cell_value is not None and str(cell_value).strip():
                                header_parts.append(str(cell_value).strip())
                    
                    # Combine header parts intelligently
                    combined_header = self.combine_header_parts(header_parts, col_idx)
                    combined_headers.append(combined_header)
                except Exception as e:
                    print(f"Error processing column {col_idx}: {e}")
                    combined_headers.append(f"Column_{col_idx + 1}")
            
            print(f"Header detection results:")
            print(f"  - Detected {header_row_count} header rows")
            print(f"  - Combined headers: {combined_headers}")
            
            return combined_headers
            
        except Exception as e:
            print(f"Critical error in header detection: {e}")
            # Ultimate fallback: create simple column names
            try:
                num_cols = len(all_rows[0]) if all_rows and len(all_rows) > 0 else 1
                return [f"Column_{i+1}" for i in range(num_cols)]
            except:
                return ["Column_1"]  # Absolute fallback
            
            # Combine header parts intelligently
            combined_header = self.combine_header_parts(header_parts, col_idx)
            combined_headers.append(combined_header)
        print(f"Header detection results:")
        print(f"  - Detected {header_row_count} header rows")
        print(f"  - Combined headers: {combined_headers}")
        
        # Validate header detection results
        validation_results = self.validate_header_detection(combined_headers, header_analysis)
        for warning in validation_results["warnings"]:
            print(f"  - Warning: {warning}")
        for recommendation in validation_results["recommendations"]:
            print(f"  - Recommendation: {recommendation}")
        
        return combined_headers
    
    def combine_header_parts(self, header_parts, col_idx):
        """Intelligently combine header parts based on their content"""
        try:
            if not header_parts:
                return f"Column_{col_idx + 1}"
            
            # Remove duplicates while preserving order
            unique_parts = []
            for part in header_parts:
                if part not in unique_parts:
                    unique_parts.append(part)
            
            if len(unique_parts) == 1:
                return unique_parts[0]
            
            # Check for specific patterns and combine accordingly
            if len(unique_parts) == 2:
                part1, part2 = unique_parts[0], unique_parts[1]
                
                # Pattern: Month + Year (e.g., "December" + "2025")
                if self.is_month_value(part1) and self.is_year_value(part2):
                    return f"{part1} {part2}"
                
                # Pattern: Year + Month (e.g., "2025" + "December")
                if self.is_year_value(part1) and self.is_month_value(part2):
                    return f"{part2} {part1}"
                
                # Pattern: Text + Year (e.g., "Revenue" + "2025")
                if not self.is_year_value(part1) and self.is_year_value(part2):
                    return f"{part1} {part2}"
                
                # Pattern: Year + Text (e.g., "2025" + "Revenue")
                if self.is_year_value(part1) and not self.is_year_value(part2):
                    return f"{part2} {part1}"
            
            # For 3+ parts or other patterns, use intelligent joining
            if len(unique_parts) >= 3:
                # Sort parts: put years at the end, months in middle, other text first
                text_parts = [p for p in unique_parts if not self.is_year_value(p) and not self.is_month_value(p)]
                month_parts = [p for p in unique_parts if self.is_month_value(p)]
                year_parts = [p for p in unique_parts if self.is_year_value(p)]
                
                # Combine in logical order: text, month, year
                ordered_parts = text_parts + month_parts + year_parts
                return " ".join(ordered_parts) if ordered_parts else " - ".join(unique_parts)
            
            # Default: join with space for natural combinations, dash for unclear relationships
            if all(len(str(part)) <= 4 and (str(part).isdigit() or self.is_month_value(part)) for part in unique_parts):
                return " ".join(unique_parts)  # For short, related parts like months/years
            else:
                return " - ".join(unique_parts)  # For longer or unrelated parts
                
        except Exception as e:
            print(f"Error combining header parts {header_parts}: {e}")
            # Fallback: simple join
            return " - ".join(str(part) for part in header_parts) if header_parts else f"Column_{col_idx + 1}"

    def determine_header_row_count(self, header_analysis):
        """Determine how many rows contain header information"""
        if not header_analysis:
            return 1
        
        # Rules for determining header rows:
        # 1. First row is usually a header
        # 2. Second row might be sub-headers (years, months, categories)
        # 3. Third row might be additional sub-headers
        # 4. Stop when we hit a row that looks like data
        
        header_count = 1  # At least one header row
        
        for i, analysis in enumerate(header_analysis):
            if i == 0:
                # First row is almost always a header
                continue
            
            try:
                current_row = analysis
                prev_row = header_analysis[i-1] if i > 0 else None
                
                # Ensure all required keys exist with safe defaults
                current_row_safe = {
                    'has_text': current_row.get('has_text', True),
                    'has_years': current_row.get('has_years', False),
                    'has_months': current_row.get('has_months', False),
                    'has_numbers_only': current_row.get('has_numbers_only', False),
                    'non_empty_count': current_row.get('non_empty_count', 0)
                }
                
                prev_row_safe = {}
                if prev_row:
                    prev_row_safe = {
                        'has_text': prev_row.get('has_text', True),
                        'has_years': prev_row.get('has_years', False),
                        'has_months': prev_row.get('has_months', False),
                        'non_empty_count': prev_row.get('non_empty_count', 0)
                    }
                
                # Check if this row looks like a continuation of headers
                is_header_continuation = False
                
                # Case 1: Row has text (likely header continuation)
                if current_row_safe['has_text'] and current_row_safe['non_empty_count'] > 0:
                    is_header_continuation = True
                
                # Case 2: Row has only years (common sub-header pattern)
                elif current_row_safe['has_years'] and not current_row_safe['has_text']:
                    is_header_continuation = True
                    print(f"  - Row {i} detected as year sub-headers")
                
                # Case 3: Row has only months (common sub-header pattern)
                elif current_row_safe['has_months'] and not current_row_safe['has_text']:
                    is_header_continuation = True
                    print(f"  - Row {i} detected as month sub-headers")
                
                # Case 4: Row has numbers/dates that could be sub-headers
                elif (current_row_safe['has_numbers_only'] and 
                      prev_row_safe and prev_row_safe['has_text'] and
                      current_row_safe['non_empty_count'] <= prev_row_safe['non_empty_count']):
                    is_header_continuation = True
                    print(f"  - Row {i} detected as numeric sub-headers")
                
                # Additional check: if previous row had months/years and current has the complement
                if (prev_row_safe and 
                    ((prev_row_safe.get('has_months') and current_row_safe['has_years']) or
                     (prev_row_safe.get('has_years') and current_row_safe['has_months']))):
                    is_header_continuation = True
                    print(f"  - Row {i} detected as complementary month/year headers")
                
                if is_header_continuation and i < 3:  # Max 3 header rows
                    header_count = i + 1
                    print(f"  - Including row {i} as header (total header rows: {header_count})")
                else:
                    print(f"  - Row {i} appears to be data, stopping header detection")
                    break
                    
            except Exception as e:
                print(f"Error processing header row {i}: {e}")
                # If we can't analyze this row, assume it's data and stop
                break
        
        return header_count

    def find_data_start_row(self, all_rows, max_header_rows=3):
        """Find the row index where actual data starts"""
        if len(all_rows) <= 1:
            return 1
        
        # Use the header detection to find where data starts
        header_analysis = []
        for row_idx in range(min(max_header_rows, len(all_rows))):
            row_values = [cell.value for cell in all_rows[row_idx]]
            row_analysis = {
                'row_index': row_idx,
                'values': row_values,
                'non_empty_count': sum(1 for val in row_values if val is not None and str(val).strip()),
                'has_text': any(isinstance(val, str) and not str(val).replace('.', '').replace('-', '').isdigit() 
                               for val in row_values if val is not None),
                'has_mixed_data': self.has_mixed_data_pattern(row_values)
            }
            header_analysis.append(row_analysis)
        
        header_row_count = self.determine_header_row_count(header_analysis)
        
        # Data starts after the header rows
        return header_row_count

    def has_mixed_data_pattern(self, row_values):
        """Check if a row has a mixed pattern typical of data rows"""
        non_empty_values = [val for val in row_values if val is not None and str(val).strip()]
        
        if len(non_empty_values) == 0:
            return False
        
        # Check for typical data patterns
        has_text = any(isinstance(val, str) and len(str(val)) > 3 for val in non_empty_values)
        has_numbers = any(self.is_numeric_or_date(val) for val in non_empty_values)
        
        # Data rows typically have both text and numbers, or are mostly numbers
        return has_text and has_numbers
    
    def is_numeric_or_date(self, value):
        """Check if a value is numeric or a date"""
        if value is None:
            return False
        
        # Check if it's already a number
        if isinstance(value, (int, float)):
            return True
        
        # Check if it's a datetime
        from datetime import datetime, date
        if isinstance(value, (datetime, date)):
            return True
        
        # Check if string represents a number or year
        str_val = str(value).strip()
        if str_val.isdigit():
            # Could be a year if it's 4 digits
            return True
        
        # Try to convert to float
        try:
            float(str_val)
            return True
        except (ValueError, TypeError):
            pass
        
        # Check for common date patterns
        import re
        date_patterns = [
            r'^\d{4}$',  # Year like 2024, 2025
            r'^\d{1,2}/\d{1,2}/\d{4}$',  # MM/DD/YYYY
            r'^\d{4}-\d{1,2}-\d{1,2}$',  # YYYY-MM-DD
        ]
        
        for pattern in date_patterns:
            if re.match(pattern, str_val):
                return True
        
        return False

    def is_year_value(self, value):
        """Check if a value represents a year (1900-2100) or could be a year-like identifier"""
        if value is None:
            return False
        
        try:
            str_val = str(value).strip()
            
            # Check if it's a 4-digit number in reasonable year range
            if str_val.isdigit() and len(str_val) == 4:
                year = int(str_val)
                return 1900 <= year <= 2100
            
            # Check if it contains a 4-digit year pattern (like "2024Q1", "FY2024", etc.)
            import re
            year_pattern = r'(19|20)\d{2}'
            matches = re.findall(year_pattern, str_val)
            if matches:
                return True
                        
            return False
        except Exception:
            return False
    
    def is_month_value(self, value):
        """Check if a value represents a month name or abbreviation"""
        if value is None:
            return False
        
        try:
            str_val = str(value).strip().lower()
            
            # Full month names
            full_months = [
                'january', 'february', 'march', 'april', 'may', 'june',
                'july', 'august', 'september', 'october', 'november', 'december'
            ]
            
            # Month abbreviations
            month_abbrevs = [
                'jan', 'feb', 'mar', 'apr', 'may', 'jun',
                'jul', 'aug', 'sep', 'oct', 'nov', 'dec'
            ]
            
            # Check for numeric months (1-12)
            if str_val.isdigit():
                month_num = int(str_val)
                return 1 <= month_num <= 12
                
            return str_val in full_months or str_val in month_abbrevs
        except Exception:
            return False

    def create_structured_results(self, file1_data, file2_data, comparison_results, primary_column):
        """Create structured results that can be easily modified from frontend"""
        
        # Get comparison data
        added_rows = comparison_results.get("rows_added", [])
        removed_rows = comparison_results.get("rows_removed", [])
        changed_values = comparison_results.get("value_diff", [])
        
        # Create lookup dictionaries for quick access
        file1_rows = {row["excel_row"]: row for row in file1_data["rows"]}
        file2_rows = {row["excel_row"]: row for row in file2_data["rows"]}
        
        # Mark added rows (exist in file2 but not in file1)
        for added_row in added_rows:
            if "_OriginalRow" in added_row:
                excel_row = added_row["_OriginalRow"]
                if excel_row in file2_rows:
                    file2_rows[excel_row]["match_status"] = "added"
        
        # Mark removed rows (exist in file1 but not in file2)
        for removed_row in removed_rows:
            if "_OriginalRow" in removed_row:
                excel_row = removed_row["_OriginalRow"]
                if excel_row in file1_rows:
                    file1_rows[excel_row]["match_status"] = "removed"
        
        # Mark changed values
        for change in changed_values:
            # File1 changes
            if "excel_row_file1" in change:
                excel_row1 = change["excel_row_file1"]
                if excel_row1 in file1_rows:
                    file1_rows[excel_row1]["match_status"] = "changed"
                    file1_rows[excel_row1]["has_changes"] = True
                    
                    # Mark specific changed columns
                    if "column" in change:
                        column_name = change["column"]
                        if column_name in file1_rows[excel_row1]["data"]:
                            file1_rows[excel_row1]["data"][column_name]["is_changed"] = True
                            file1_rows[excel_row1]["changed_columns"].append(column_name)
            
            # File2 changes
            if "excel_row_file2" in change:
                excel_row2 = change["excel_row_file2"]
                if excel_row2 in file2_rows:
                    file2_rows[excel_row2]["match_status"] = "changed"
                    file2_rows[excel_row2]["has_changes"] = True
                    
                    # Mark specific changed columns
                    if "column" in change:
                        column_name = change["column"]
                        if column_name in file2_rows[excel_row2]["data"]:
                            file2_rows[excel_row2]["data"][column_name]["is_changed"] = True
                            file2_rows[excel_row2]["changed_columns"].append(column_name)
        
        # Mark remaining rows as matched
        for row in file1_data["rows"]:
            if row["match_status"] == "pending":
                row["match_status"] = "matched"
        
        for row in file2_data["rows"]:
            if row["match_status"] == "pending":
                row["match_status"] = "matched"
        
        # Calculate statistics
        file1_stats = self.calculate_statistics(file1_data["rows"])
        file2_stats = self.calculate_statistics(file2_data["rows"])
        
        return {
            "comparison_metadata": {
                "primary_column": primary_column,
                "timestamp": datetime.now().isoformat(),
                "ranges": {
                    "file1": file1_data["range"],
                    "file2": file2_data["range"]
                }
            },
            "file1": {
                "headers": file1_data["headers"],
                "rows": file1_data["rows"],
                "statistics": file1_stats,
                "total_rows": file1_data["total_rows"]
            },
            "file2": {
                "headers": file2_data["headers"],
                "rows": file2_data["rows"],
                "statistics": file2_stats,
                "total_rows": file2_data["total_rows"]
            },
            "detailed_changes": changed_values,
            "original_comparison": comparison_results
        }

    def analyze_header_structure(self, all_rows, max_header_rows=3):
        """Analyze and provide detailed information about header structure"""
        if len(all_rows) == 0:
            return {
                "detected_pattern": "no_data",
                "header_rows": 0,
                "confidence": "low",
                "description": "No data found in the range"
            }
        
        # Analyze each potential header row
        analysis_details = []
        for row_idx in range(min(max_header_rows, len(all_rows))):
            row_values = [cell.value for cell in all_rows[row_idx]]
            
            # Count different types of values
            text_values = [v for v in row_values if v is not None and isinstance(v, str) and not str(v).isdigit()]
            year_values = [v for v in row_values if self.is_year_value(v)]
            month_values = [v for v in row_values if self.is_month_value(v)]
            numeric_values = [v for v in row_values if self.is_numeric_or_date(v) and not self.is_year_value(v)]
            empty_values = [v for v in row_values if v is None or str(v).strip() == '']
            
            analysis_details.append({
                "row_index": row_idx,
                "total_cells": len(row_values),
                "text_count": len(text_values),
                "year_count": len(year_values),
                "month_count": len(month_values),
                "numeric_count": len(numeric_values),
                "empty_count": len(empty_values),
                "sample_values": [str(v) for v in row_values[:5] if v is not None]
            })
        
        # Determine header pattern
        if len(analysis_details) == 0:
            pattern = "no_data"
            confidence = "low"
        elif len(analysis_details) == 1:
            pattern = "single_row"
            confidence = "high"
        else:
            # Analyze multi-row patterns
            first_row = analysis_details[0]
            second_row = analysis_details[1] if len(analysis_details) > 1 else None
            
            if second_row:
                # Month + Year pattern
                if first_row["month_count"] > 0 and second_row["year_count"] > 0:
                    pattern = "month_year"
                    confidence = "high"
                # Text + Year pattern
                elif first_row["text_count"] > 0 and second_row["year_count"] > 0:
                    pattern = "text_year"
                    confidence = "high"
                # Year + Month pattern
                elif first_row["year_count"] > 0 and second_row["month_count"] > 0:
                    pattern = "year_month"
                    confidence = "high"
                # General text continuation
                elif first_row["text_count"] > 0 and second_row["text_count"] > 0:
                    pattern = "multi_text"
                    confidence = "medium"
                # Mixed pattern
                else:
                    pattern = "mixed"
                    confidence = "medium"
            else:
                pattern = "single_row"
                confidence = "high"
        
        # Generate description
        descriptions = {
            "no_data": "No data found in the specified range",
            "single_row": "Single row of headers detected",
            "month_year": "Month names in first row, years in second row",
            "text_year": "Text headers in first row, years in second row",
            "year_month": "Years in first row, month names in second row",
            "multi_text": "Multiple rows of text headers",
            "mixed": "Mixed header pattern detected"
        }
        
        return {
            "detected_pattern": pattern,
            "header_rows": len([a for a in analysis_details if a["text_count"] > 0 or a["year_count"] > 0 or a["month_count"] > 0]),
            "confidence": confidence,
            "description": descriptions.get(pattern, "Unknown pattern"),
            "analysis_details": analysis_details
        }

    def validate_header_detection(self, headers, analysis_details):
        """Validate the header detection results and provide warnings if needed"""
        warnings = []
        recommendations = []
        
        # Check for duplicate headers
        duplicate_headers = []
        seen_headers = set()
        for header in headers:
            if header in seen_headers:
                duplicate_headers.append(header)
            seen_headers.add(header)
        
        if duplicate_headers:
            warnings.append(f"Duplicate headers detected: {', '.join(duplicate_headers)}")
            recommendations.append("Consider adjusting the range to include more specific header rows")
        
        # Check for generic column names
        generic_count = len([h for h in headers if h.startswith('Column_')])
        if generic_count > len(headers) * 0.3:  # More than 30% generic
            warnings.append(f"{generic_count} columns have generic names (Column_X)")
            recommendations.append("The range might not include proper header rows")
        
        # Check for very short headers that might be incomplete
        short_headers = [h for h in headers if len(h.strip()) <= 2 and not h.strip().isdigit()]
        if short_headers:
            warnings.append(f"Very short headers detected: {', '.join(short_headers)}")
            recommendations.append("Headers might be incomplete - consider including more header rows")
        
        # Check for headers that are only numbers (might be data)
        numeric_headers = [h for h in headers if h.strip().isdigit()]
        if len(numeric_headers) > len(headers) * 0.5:  # More than 50% numeric
            warnings.append("More than half of headers are numbers - this might be data instead of headers")
            recommendations.append("Check if the range starts too low and includes data rows as headers")
        
        return {
            "warnings": warnings,
            "recommendations": recommendations,
            "quality_score": max(0, 100 - len(warnings) * 20)  # 100 = perfect, decreases by 20 per warning
        }

    def calculate_statistics(self, rows):
        """Calculate match statistics for a set of rows"""
        total = len(rows)
        matched = len([r for r in rows if r["match_status"] == "matched"])
        changed = len([r for r in rows if r["match_status"] == "changed"])
        added = len([r for r in rows if r["match_status"] == "added"])
        removed = len([r for r in rows if r["match_status"] == "removed"])
        
        return {
            "total_rows": total,
            "matched": matched,
            "changed": changed,
            "added": added,
            "removed": removed,
            "match_percentage": round((matched / total * 100), 2) if total > 0 else 0,
            "change_percentage": round(((changed + added + removed) / total * 100), 2) if total > 0 else 0
        }


class DeleteComparisonView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, comparison_id):
        comparison = Comparison.objects.filter(id=comparison_id).first()
        if not comparison:
            return Response({"message": "Comparison not found."}, status=status.HTTP_404_NOT_FOUND)

        comparison.delete()
        return Response({"message": "Comparison deleted successfully."}, status=status.HTTP_200_OK)


class ListUserComparisonView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        comparisons = Comparison.objects.filter(user=request.user)
    
        search_query = request.GET.get('search', '').strip()
        if search_query:
            comparisons = comparisons.filter(
                Q(file1__file__icontains=search_query) |
                Q(file2__file__icontains=search_query)
            )

        comparison_id = request.GET.get('comparison_id', '').strip()
        if comparison_id:
            comparisons = comparisons.filter(id=comparison_id)

        # Date filtering
        date_filter_type = request.GET.get('date_filter_type', '')
        if date_filter_type == 'range':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            if start_date and end_date:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                end_date = end_date_obj.strftime('%Y-%m-%d')
                comparisons = comparisons.filter(timestamp__range=[start_date, end_date])
        elif date_filter_type == 'day' and request.GET.get('filter_date'):
            filter_date = request.GET.get('filter_date')
            comparisons = comparisons.filter(timestamp__date=filter_date)
        elif date_filter_type == 'month' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET.get('filter_date'), '%Y-%m-%d')
            comparisons = comparisons.filter(
                timestamp__year=filter_date.year,
                timestamp__month=filter_date.month
            )
        elif date_filter_type == 'year' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET.get('filter_date'), '%Y-%m-%d')
            comparisons = comparisons.filter(timestamp__year=filter_date.year)

        paginator = PageNumberPagination()
        paginator.page_size = 10
        results_page = paginator.paginate_queryset(comparisons, request)
        serializer = ComparisonSerializer(results_page, many=True)
        data = serializer.data
        for i, comp in enumerate(results_page):
            data[i]['file1_name'] = comp.file1.file.name
            data[i]['file2_name'] = comp.file2.file.name
            data[i]['user'] = comp.user.username
        return Response({
            'results': data,
            'current_page': paginator.page.number,
            'total_pages': paginator.page.paginator.num_pages
        }, status=status.HTTP_200_OK)


class ListAllComparisonView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        comparisons = Comparison.objects.all().select_related('file1', 'file2', 'user')

        search_query = request.GET.get('search', '').strip()
        if search_query:
            comparisons = comparisons.filter(
                Q(file1__file__icontains=search_query) |
                Q(file2__file__icontains=search_query)
            )

        comparison_id = request.GET.get('comparison_id', '').strip()
        if comparison_id:
            comparisons = comparisons.filter(id=comparison_id)

        # Date filtering
        date_filter_type = request.GET.get('date_filter_type', '')
        if date_filter_type == 'range':
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            if start_date and end_date:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                comparisons = comparisons.filter(timestamp__range=[start_date, end_date_obj.strftime('%Y-%m-%d')])
        elif date_filter_type == 'day' and request.GET.get('filter_date'):
            comparisons = comparisons.filter(timestamp__date=request.GET['filter_date'])
        elif date_filter_type == 'month' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET['filter_date'], '%Y-%m-%d')
            comparisons = comparisons.filter(timestamp__year=filter_date.year, timestamp__month=filter_date.month)
        elif date_filter_type == 'year' and request.GET.get('filter_date'):
            filter_date = datetime.strptime(request.GET['filter_date'], '%Y-%m-%d')
            comparisons = comparisons.filter(timestamp__year=filter_date.year)

        comparisons = comparisons.order_by('user__id', '-timestamp')

        paginator = PageNumberPagination()
        paginator.page_size = 10
        results_page = paginator.paginate_queryset(comparisons, request)
        data = []
        for comp in results_page:
            data.append({
                'id': comp.id,
                'file1_name': comp.file1.file.name,
                'file2_name': comp.file2.file.name,
                'user': comp.user.username,
                'timestamp': comp.timestamp
            })
        return paginator.get_paginated_response(data)


class ComparisonDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, comparison_id):
        if request.user.role == 'admin':
            comparison = Comparison.objects.filter(id=comparison_id).first()
        else:
            comparison = Comparison.objects.filter(id=comparison_id, user=request.user).first()

        if not comparison:
            return Response({"error": "Comparison not found or you don't have permission to view it"}, 
                           status=status.HTTP_404_NOT_FOUND)

        # Parse results to check if it's structured format
        if isinstance(comparison.results, str):
            try:
                results_data = json.loads(comparison.results)
            except json.JSONDecodeError:
                results_data = comparison.results
        else:
            results_data = comparison.results

        # Check if results are in new structured format
        is_structured = (
            isinstance(results_data, dict) and 
            "comparison_metadata" in results_data and 
            "file1" in results_data and 
            "file2" in results_data
        )

        data = {
            'id': comparison.id,
            'file1_name': comparison.file1.file.name,
            'file2_name': comparison.file2.file.name,
            'user': comparison.user.username,
            'timestamp': comparison.timestamp,
            'range1': comparison.range1,
            'range2': comparison.range2,
            'column1': comparison.column1,
            'column2': comparison.column2,
            'results': results_data,
            'is_structured_format': is_structured
        }
        
        if is_structured:
            # Add summary statistics for frontend
            file1_stats = results_data.get("file1", {}).get("statistics", {})
            file2_stats = results_data.get("file2", {}).get("statistics", {})
            
            data['summary'] = {
                'file1_stats': file1_stats,
                'file2_stats': file2_stats,
                'total_file1_rows': results_data.get("file1", {}).get("total_rows", 0),
                'total_file2_rows': results_data.get("file2", {}).get("total_rows", 0)
            }
        
        return Response(data, status=status.HTTP_200_OK)


def column_letter_to_index(col_letter):
    """Convert column letter (A, B, C, etc.) to 0-based index"""
    result = 0
    for char in col_letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1
def get_excel_column_name(column_index):
    """Convert 0-based column index to Excel column letter(s)"""
    result = ""
    while column_index >= 0:
        result = chr(column_index % 26 + ord('A')) + result
        column_index = column_index // 26 - 1
        if column_index < 0:
            break
    return result
def get_next_column_letter(col_letter):
    """Get the next column letter (A->B, Z->AA, etc.)"""
    col_letter = col_letter.upper()
    
    # Convert to list for easier manipulation
    letters = list(col_letter)
    
    # Start from the rightmost character
    i = len(letters) - 1
    
    while i >= 0:
        if letters[i] != 'Z':
            # Simple case: increment the letter
            letters[i] = chr(ord(letters[i]) + 1)
            break
        else:
            # This letter is Z, so it becomes A and we carry over
            letters[i] = 'A'
            if i == 0:
                # We've carried over past the first letter, so add a new A at the beginning
                letters.insert(0, 'A')
                break
            # Continue to the next letter to the left
            i -= 1
    
    return ''.join(letters)
def safe_assign_to_dataframe(df, row_index, column_index, value, column_name=None):
    """
    Safely assign a value to df.iloc[row_index, column_index].
    Expands the DataFrame if necessary.
    """
    # Ensure we have enough rows
    if row_index >= len(df):
        # Add missing rows
        missing_rows = row_index - len(df) + 1
        for _ in range(missing_rows):
            df.loc[len(df)] = [None] * len(df.columns)
    
    # Ensure we have enough columns
    current_columns = len(df.columns)
    if column_index >= current_columns:
        # Add missing columns
        for i in range(current_columns, column_index + 1):
            if column_name and i == column_index:
                col_name = column_name
            else:
                col_name = get_excel_column_name(i)
            df[col_name] = None
    
    # Now safely assign the value
    df.iloc[row_index, column_index] = value
    return df




class DownloadComparisonExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, comparison_id, file_number=1):
        """
        Download the Excel file from a comparison.
        file_number: 1 for first file, 2 for second file
        """
        try:
            # For admin, allow access to any comparison
            if request.user.role == 'admin':
                comparison = Comparison.objects.get(id=comparison_id)
            else:
                # For regular users, only allow access to their own comparisons
                comparison = Comparison.objects.get(id=comparison_id, user=request.user)
                
            # Determine which file to download
            if file_number == 1:
                file_obj = comparison.file1
                file_label = "source"
            elif file_number == 2:
                file_obj = comparison.file2
                file_label = "target"
            else:
                return Response({
                    'message': 'Invalid file number. Use 1 for source file or 2 for target file.',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get the file path and comparison data
            range1 = comparison.range1
            range2 = comparison.range2
            result = comparison.results
            
            if isinstance(result, str):
                result_data = json.loads(result)
            else:
                result_data = result
                
            print(f"Range1: {range1}, Range2: {range2}")
            print(f"Comparison Results: {result}")
            
            file_path = file_obj.file.path
            
            # Check if file exists
            if not os.path.exists(file_path):
                return Response({
                    'message': f'File not found: {file_label} file for comparison {comparison_id}',
                    'status': 'error'
                }, status=status.HTTP_404_NOT_FOUND)
                
            # Decrypt the file temporarily for download
            try:
                # Check the file size before decryption
                file_size = os.path.getsize(file_path)
                print(f"Original file size: {file_size} bytes")
                
                # First, try to read the file directly to see if it's already decrypted
                with open(file_path, 'rb') as f:
                    first_bytes = f.read(10)
                    print(f"First 10 bytes of file: {first_bytes}")
                
                # Check if the file is already in Excel format (not encrypted)
                if first_bytes.startswith(b'PK') or first_bytes.startswith(b'\xd0\xcf\x11\xe0'):
                    print("File appears to be unencrypted Excel format")
                    # Read the entire file as-is
                    with open(file_path, 'rb') as f:
                        decrypted_content = f.read()
                    print(f"Direct file read size: {len(decrypted_content)} bytes")
                else:
                    print("File appears to be encrypted, attempting decryption...")
                    # Try to decrypt the file
                    try:
                        decrypted_result = decrypt_file(file_path)
                        print(f"Decrypt function returned: type={type(decrypted_result)}, length={len(str(decrypted_result))}")
                        
                        # Check if decrypt_file returned a path (string) instead of content
                        if isinstance(decrypted_result, str):
                            if os.path.exists(decrypted_result):
                                print(f"Decrypt function returned a file path: {decrypted_result}")
                                print("Reading the decrypted file content...")
                                with open(decrypted_result, 'rb') as f:
                                    decrypted_content = f.read()
                                print(f"File content read successfully: {len(decrypted_content)} bytes")
                            else:
                                print("Decrypt function returned a string, but it's not a valid path")
                                print(f"First 100 characters: {decrypted_result[:100]}")
                                # Try to encode the string as bytes
                                decrypted_content = decrypted_result.encode('latin-1')
                        else:
                            # Decrypt function returned bytes directly
                            decrypted_content = decrypted_result
                            
                        print(f"Final decrypted content size: {len(decrypted_content)} bytes")
                        print(f"Final decrypted content type: {type(decrypted_content)}")
                        
                    except Exception as decrypt_error:
                        print(f"Decryption failed: {str(decrypt_error)}")
                        return Response({
                            'message': f'Error decrypting file: {str(decrypt_error)}',
                            'status': 'error'
                        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
                # Ensure decrypted_content is bytes
                if isinstance(decrypted_content, str):
                    print("Converting string to bytes using latin-1 encoding")
                    try:
                        decrypted_content = decrypted_content.encode('latin-1')
                    except UnicodeEncodeError:
                        print("Failed to encode with latin-1, trying utf-8")
                        try:
                            decrypted_content = decrypted_content.encode('utf-8')
                        except UnicodeEncodeError:
                            print("Both encoding methods failed, file might be corrupted")
                            return Response({
                                'message': 'File encoding error - file may be corrupted',
                                'status': 'error'
                            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
                # Verify we have bytes now
                if not isinstance(decrypted_content, bytes):
                    return Response({
                        'message': f'Error: Decrypted content is {type(decrypted_content)}, expected bytes',
                        'status': 'error'
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                    
                # Verify the decrypted content starts with valid Excel magic bytes
                if decrypted_content.startswith(b'PK'):
                    print("File appears to be a valid zip-based format (Excel .xlsx)")
                elif decrypted_content.startswith(b'\xd0\xcf\x11\xe0'):
                    print("File appears to be OLE format (Excel .xls)")
                else:
                    print(f"Warning: File doesn't start with expected Excel headers. First 10 bytes: {decrypted_content[:10]}")
                    print("File might still be valid - proceeding with download")
                    
                # Modify the existing Excel file structure
                try:
                    # Create a BytesIO object from the decrypted content
                    excel_buffer = io.BytesIO(decrypted_content)
                    
                    # Load the workbook using openpyxl to preserve original structure
                    workbook = load_workbook(excel_buffer)
                    
                    # Get the first worksheet (or specify sheet name if needed)
                    worksheet = workbook.active  # or workbook['SheetName']
                    
                    # Parse the range
                    rangee = range2.split(':')
                    Range1 = rangee[0].strip().upper()
                    Range2 = rangee[1].strip().upper() if len(rangee) > 1 else None

                    match = re.match(r'([A-Z]+)(\d+)', Range1)
                    match2 = re.match(r'([A-Z]+)(\d+)', Range2)
                    
                    if match and match2:
                        col_letter = match.group(1)
                        row_number = int(match.group(2))
                        col_letter2 = match2.group(1)
                        row_number2 = int(match2.group(2))
                        next_column = get_next_column_letter(col_letter2)
                        
                        print(f"Current column: {col_letter2}, Next column: {next_column}")

                        col_letter_index = column_letter_to_index(col_letter)
                        next_column_index = column_letter_to_index(next_column)
                        
                        # Get comparison results
                        added_row = result_data.get("rows_added", [])
                        removed_row = result_data.get("rows_removed", [])
                        changed_values = result_data.get("value_diff", [])
                        column= comparison.column1
                        
                        # Create a set of rows to mark as unmatched for efficiency
                        unmatched_rows = set()
                        workbook2 = load_workbook(excel_buffer)
                        worksheet2 = workbook2.active
                        store_matched_rows = []

                        for row_num in range(row_number+2, row_number2): # +1 to skip header row
                            cell_value = worksheet2.cell(row=row_num, column=col_letter_index+1).value
                            if cell_value is None or cell_value == '':
                                continue
                            store_matched_rows.append(row_num)
                            print(cell_value) 
                        
                        # Process added rows
                        for item in added_row:                            
                              if column in item and column in item and item[column] is not None:
                                unmatched_rows.add(item['_OriginalRow'])
                        

                        # Process changed values
                        for item in changed_values:
                            if file_number == 1 and 'excel_row_file1' in item:
                                unmatched_rows.add(item['excel_row_file1'])
                            elif file_number == 2 and 'excel_row_file2' in item:
                                unmatched_rows.add(item['excel_row_file2'])
                        
                        # Add header for the new column if it doesn't exist
                        header_cell = worksheet.cell(row=row_number, column=next_column_index + 1)
                        if header_cell.value is None or header_cell.value == '':
                            header_cell.value = "Comparison Status"
                            header_cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Yellow

                        

                        total_matched = 0
                        total_unmatched = len(unmatched_rows)   
                        # Mark unmatched rows in the new column
                        for row_target in range(row_number + 2, row_number2 ):  # +1 to skip header row
                            cell = worksheet.cell(row=row_target, column=next_column_index + 1)
                            
                            if row_target in unmatched_rows:
                                cell.value = "unmatched"
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
                                print(f"Row {row_target} in column {next_column} marked as unmatched")

                            elif row_target in store_matched_rows:
                                cell.value = "match"
                                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                                total_matched += 1
                                print(f"Row {row_target} in column {next_column} marked as match")

                        total_rows_with_data = len(store_matched_rows)
                        matched_percentage = (total_matched / total_rows_with_data * 100) if total_rows_with_data > 0 else 0
                        unmatched_percentage = (total_unmatched / total_rows_with_data * 100) if total_rows_with_data > 0 else 0

                        cell=worksheet.cell(row=row_number , column=next_column_index + 2)
                        cell.value = "Total matched"
                        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Medium gray for header



                        cell = worksheet.cell(row=row_number + 3, column=next_column_index + 2)
                        cell.value = f"{total_matched} ({matched_percentage:.1f}%)"
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                          # Light gray

                        cell = worksheet.cell(row=row_number, column=next_column_index + 3)
                        cell.value = "Total unmatched"
                        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

                        cell = worksheet.cell(row=row_number + 3, column=next_column_index + 3)
                        cell.value = f"{total_unmatched} ({unmatched_percentage:.1f}%)"
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow

                        comparison_status_col = get_excel_column_name(next_column_index)
                        worksheet.column_dimensions[comparison_status_col].width = 20

                        comparison_status_col = get_excel_column_name(next_column_index+1)
                        worksheet.column_dimensions[comparison_status_col].width = 20

                        comparison_status_col = get_excel_column_name(next_column_index+2)
                        worksheet.column_dimensions[comparison_status_col].width = 20
                            
                    
                    # Save the modified workbook to a new BytesIO buffer
                    modified_excel_buffer = io.BytesIO()
                    workbook.save(modified_excel_buffer)
                    
                    # Get the modified content
                    modified_excel_buffer.seek(0)
                    final_content = modified_excel_buffer.getvalue()
                    
                    print(f"Modified Excel content size: {len(final_content)} bytes")
                    
                except Exception as excel_modify_error:
                    print(f"Error modifying Excel file: {str(excel_modify_error)}")
                    # Fallback to original content if modification fails
                    final_content = decrypted_content
                
                # Get the original filename without path
                original_filename = os.path.basename(file_obj.file.name)
                
                # Ensure proper Excel content type
                if original_filename.lower().endswith('.xlsx'):
                    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                elif original_filename.lower().endswith('.xls'):
                    content_type = 'application/vnd.ms-excel'
                elif original_filename.lower().endswith('.csv'):
                    content_type = 'text/csv'
                else:
                    # Default to Excel format
                    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    
                # Create the response with the modified content
                response = HttpResponse(final_content, content_type=content_type)
                response['Content-Disposition'] = f'attachment; filename="{original_filename}"'
                response['Content-Length'] = len(final_content)
                response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                response['Pragma'] = 'no-cache'
                response['Expires'] = '0'
                
                print(f"Sending response with {len(final_content)} bytes, content-type: {content_type}")
                return response
                
            except Exception as e:
                return Response({
                    'message': f'Error processing file: {str(e)}',
                    'status': 'error'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Comparison.DoesNotExist:
            return Response({
                'message': 'Comparison not found or you do not have permission to access it',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'message': f'Error: {str(e)}',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AnalyzeExcelTableView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, comparison_id, file_number=1):
        """Analyze Excel table structure and return the first empty cell after the table."""
        try:
            if request.user.role == 'admin':
                comparison = Comparison.objects.get(id=comparison_id)
            else:
                comparison = Comparison.objects.get(id=comparison_id, user=request.user)
            
            if file_number == 1:
                file_obj = comparison.file1
                file_label = "source"
            elif file_number == 2:
                file_obj = comparison.file2
                file_label = "target"
            else:
                return Response({
                    'message': 'Invalid file number. Use 1 for source file or 2 for target file.',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            file_path = file_obj.file.path
            
            if not os.path.exists(file_path):
                return Response({
                    'message': f'File not found: {file_label} file for comparison {comparison_id}',
                    'status': 'error'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # File analysis logic would go here
            return Response({
                'status': 'success',
                'file_label': file_label,
                'filename': os.path.basename(file_obj.file.name),
                'sheet_count': 0,
                'sheets': {}
            }, status=status.HTTP_200_OK)
                
        except Comparison.DoesNotExist:
            return Response({
                'message': 'Comparison not found or you do not have permission to access it',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'message': f'Error: {str(e)}',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdateComparisonView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, comparison_id):
        """Update comparison results with frontend modifications"""
        try:
            if request.user.role == 'admin':
                comparison = Comparison.objects.get(id=comparison_id)
            else:
                comparison = Comparison.objects.get(id=comparison_id, user=request.user)
            
            updated_data = request.data.get('comparison_data')
            if not updated_data:
                return Response(
                    {"error": "comparison_data is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            # Validate the structure
            if not self.validate_comparison_data(updated_data):
                return Response(
                    {"error": "Invalid comparison data structure"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            # Update timestamp
            updated_data["comparison_metadata"]["last_modified"] = datetime.now().isoformat()
            updated_data["comparison_metadata"]["modified_by"] = request.user.username
            
            # Recalculate statistics
            updated_data["file1"]["statistics"] = self.calculate_statistics(updated_data["file1"]["rows"])
            updated_data["file2"]["statistics"] = self.calculate_statistics(updated_data["file2"]["rows"])
            
            # Save updated results
            comparison.results = json.dumps(updated_data)
            comparison.save()
            
            return Response({
                "message": "Comparison updated successfully",
                "data": updated_data
            }, status=status.HTTP_200_OK)
            
        except Comparison.DoesNotExist:
            return Response(
                {"error": "Comparison not found or you don't have permission to modify it"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"error": f"Error updating comparison: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
    
    def validate_comparison_data(self, data):
        """Validate the structure of comparison data"""
        required_keys = ["comparison_metadata", "file1", "file2"]
        if not all(key in data for key in required_keys):
            return False
        
        # Validate file data structure
        for file_key in ["file1", "file2"]:
            file_data = data[file_key]
            if not all(key in file_data for key in ["headers", "rows", "total_rows"]):
                return False
            
            # Validate row structure
            for row in file_data["rows"]:
                required_row_keys = ["row_number", "excel_row", "source", "data", "match_status"]
                if not all(key in row for key in required_row_keys):
                    return False
        
        return True
    
    def calculate_statistics(self, rows):
        """Calculate match statistics for a set of rows"""
        total = len(rows)
        matched = len([r for r in rows if r["match_status"] == "matched"])
        changed = len([r for r in rows if r["match_status"] == "changed"])
        added = len([r for r in rows if r["match_status"] == "added"])
        removed = len([r for r in rows if r["match_status"] == "removed"])
        
        return {
            "total_rows": total,
            "matched": matched,
            "changed": changed,
            "added": added,
            "removed": removed,
            "match_percentage": round((matched / total * 100), 2) if total > 0 else 0,
            "change_percentage": round(((changed + added + removed) / total * 100), 2) if total > 0 else 0
        }


class ComparisonRowDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, comparison_id, file_number, row_id):
        """Get detailed information for a specific row in the comparison"""
        try:
            if request.user.role == 'admin':
                comparison = Comparison.objects.get(id=comparison_id)
            else:
                comparison = Comparison.objects.get(id=comparison_id, user=request.user)
            
            # Parse results
            if isinstance(comparison.results, str):
                results_data = json.loads(comparison.results)
            else:
                results_data = comparison.results
            
            # Validate structured format
            if not (isinstance(results_data, dict) and "file1" in results_data and "file2" in results_data):
                return Response(
                    {"error": "Comparison data is not in structured format"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            # Get file data
            file_key = f"file{file_number}"
            if file_key not in results_data:
                return Response(
                    {"error": f"Invalid file number: {file_number}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            file_data = results_data[file_key]
            
            # Find the specific row
            target_row = None
            for row in file_data["rows"]:
                if row["row_number"] == int(row_id):
                    target_row = row
                    break
            
            if not target_row:
                return Response(
                    {"error": f"Row {row_id} not found in file {file_number}"},
                    status=status.HTTP_404_NOT_FOUND,
                )
            
            # Find corresponding row in other file if exists
            other_file_key = "file1" if file_number == 2 else "file2"
            corresponding_row = None
            primary_column = results_data.get("comparison_metadata", {}).get("primary_column")
            
            if primary_column and primary_column in target_row["data"]:
                primary_value = target_row["data"][primary_column]["value"]
                
                for row in results_data[other_file_key]["rows"]:
                    if (primary_column in row["data"] and 
                        row["data"][primary_column]["value"] == primary_value):
                        corresponding_row = row
                        break
            
            return Response({
                "row_details": target_row,
                "corresponding_row": corresponding_row,
                "file_number": file_number,
                "headers": file_data["headers"],
                "primary_column": primary_column
            }, status=status.HTTP_200_OK)
            
        except Comparison.DoesNotExist:
            return Response(
                {"error": "Comparison not found or you don't have permission to view it"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"error": f"Error retrieving row details: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class UpdateComparisonRowView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, comparison_id, file_number, row_id):
        """Update a specific row's match status or data"""
        try:
            if request.user.role == 'admin':
                comparison = Comparison.objects.get(id=comparison_id)
            else:
                comparison = Comparison.objects.get(id=comparison_id, user=request.user)
            
            # Parse results
            if isinstance(comparison.results, str):
                results_data = json.loads(comparison.results)
            else:
                results_data = comparison.results
            
            # Validate structured format
            if not (isinstance(results_data, dict) and "file1" in results_data and "file2" in results_data):
                return Response(
                    {"error": "Comparison data is not in structured format"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            # Get update data
            new_match_status = request.data.get('match_status')
            updated_columns = request.data.get('updated_columns', {})
            
            # Validate match status
            valid_statuses = ['matched', 'changed', 'added', 'removed', 'pending']
            if new_match_status and new_match_status not in valid_statuses:
                return Response(
                    {"error": f"Invalid match status. Must be one of: {valid_statuses}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            # Get file data
            file_key = f"file{file_number}"
            if file_key not in results_data:
                return Response(
                    {"error": f"Invalid file number: {file_number}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            file_data = results_data[file_key]
            
            # Find and update the specific row
            row_updated = False
            for i, row in enumerate(file_data["rows"]):
                if row["row_number"] == int(row_id):
                    # Update match status
                    if new_match_status:
                        row["match_status"] = new_match_status
                    
                    # Update column data
                    for column_name, new_value in updated_columns.items():
                        if column_name in row["data"]:
                            old_value = row["data"][column_name]["value"]
                            row["data"][column_name]["value"] = new_value
                            
                            # Mark as changed if value is different
                            if old_value != new_value:
                                row["data"][column_name]["is_changed"] = True
                                if column_name not in row["changed_columns"]:
                                    row["changed_columns"].append(column_name)
                                row["has_changes"] = True
                            else:
                                row["data"][column_name]["is_changed"] = False
                                if column_name in row["changed_columns"]:
                                    row["changed_columns"].remove(column_name)
                    
                    # Update has_changes based on changed_columns
                    row["has_changes"] = len(row["changed_columns"]) > 0
                    
                    file_data["rows"][i] = row
                    row_updated = True
                    break
            
            if not row_updated:
                return Response(
                    {"error": f"Row {row_id} not found in file {file_number}"},
                    status=status.HTTP_404_NOT_FOUND,
                )
            
            # Recalculate statistics
            results_data[file_key]["statistics"] = self.calculate_statistics(file_data["rows"])
            
            # Update metadata
            results_data["comparison_metadata"]["last_modified"] = datetime.now().isoformat()
            results_data["comparison_metadata"]["modified_by"] = request.user.username
            
            # Save updated results
            comparison.results = json.dumps(results_data)
            comparison.save()
            
            return Response({
                "message": "Row updated successfully",
                "updated_row": next(row for row in file_data["rows"] if row["row_number"] == int(row_id)),
                "statistics": results_data[file_key]["statistics"]
            }, status=status.HTTP_200_OK)
            
        except Comparison.DoesNotExist:
            return Response(
                {"error": "Comparison not found or you don't have permission to modify it"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"error": f"Error updating row: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
    
    def calculate_statistics(self, rows):
        """Calculate match statistics for a set of rows"""
        total = len(rows)
        matched = len([r for r in rows if r["match_status"] == "matched"])
        changed = len([r for r in rows if r["match_status"] == "changed"])
        added = len([r for r in rows if r["match_status"] == "added"])
        removed = len([r for r in rows if r["match_status"] == "removed"])
        
        return {
            "total_rows": total,
            "matched": matched,
            "changed": changed,
            "added": added,
            "removed": removed,
            "match_percentage": round((matched / total * 100), 2) if total > 0 else 0,
            "change_percentage": round(((changed + added + removed) / total * 100), 2) if total > 0 else 0
        }


class ExportComparisonView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, comparison_id):
        """Export comparison results in various formats"""
        try:
            if request.user.role == 'admin':
                comparison = Comparison.objects.get(id=comparison_id)
            else:
                comparison = Comparison.objects.get(id=comparison_id, user=request.user)
            
            export_format = request.GET.get('format', 'json')  # json, csv, excel
            include_matched = request.GET.get('include_matched', 'true').lower() == 'true'
            include_unchanged = request.GET.get('include_unchanged', 'false').lower() == 'true'
            
            # Parse results
            if isinstance(comparison.results, str):
                results_data = json.loads(comparison.results)
            else:
                results_data = comparison.results
            
            # Validate structured format
            if not (isinstance(results_data, dict) and "file1" in results_data and "file2" in results_data):
                return Response(
                    {"error": "Comparison data is not in structured format"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            if export_format == 'json':
                return self.export_as_json(results_data, include_matched, include_unchanged)
            elif export_format == 'csv':
                return self.export_as_csv(results_data, include_matched, include_unchanged)
            elif export_format == 'excel':
                return self.export_as_excel(results_data, include_matched, include_unchanged, comparison)
            else:
                return Response(
                    {"error": "Invalid export format. Supported: json, csv, excel"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
                
        except Comparison.DoesNotExist:
            return Response(
                {"error": "Comparison not found or you don't have permission to access it"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"error": f"Error exporting comparison: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
    
    def filter_rows(self, rows, include_matched, include_unchanged):
        """Filter rows based on export preferences"""
        filtered_rows = []
        
        for row in rows:
            status = row["match_status"]
            
            # Skip matched rows if not included
            if status == "matched" and not include_matched:
                continue
            
            # Skip unchanged rows if not included
            if not row["has_changes"] and not include_unchanged:
                continue
            
            filtered_rows.append(row)
        
        return filtered_rows
    
    def export_as_json(self, results_data, include_matched, include_unchanged):
        """Export as JSON format"""
        export_data = {
            "metadata": results_data["comparison_metadata"],
            "file1": {
                "headers": results_data["file1"]["headers"],
                "rows": self.filter_rows(results_data["file1"]["rows"], include_matched, include_unchanged),
                "statistics": results_data["file1"]["statistics"]
            },
            "file2": {
                "headers": results_data["file2"]["headers"],
                "rows": self.filter_rows(results_data["file2"]["rows"], include_matched, include_unchanged),
                "statistics": results_data["file2"]["statistics"]
            }
        }
        
        response = JsonResponse(export_data, safe=False)
        response['Content-Disposition'] = 'attachment; filename="comparison_export.json"'
        return response
    
    def export_as_csv(self, results_data, include_matched, include_unchanged):
        """Export as CSV format"""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        headers = ["File", "Row_Number", "Excel_Row", "Match_Status", "Has_Changes"] + results_data["file1"]["headers"]
        writer.writerow(headers)
        
        # Write file1 data
        file1_rows = self.filter_rows(results_data["file1"]["rows"], include_matched, include_unchanged)
        for row in file1_rows:
            row_data = [
                "File1",
                row["row_number"],
                row["excel_row"],
                row["match_status"],
                row["has_changes"]
            ]
            
            # Add column values
            for header in results_data["file1"]["headers"]:
                if header in row["data"]:
                    row_data.append(row["data"][header]["value"])
                else:
                    row_data.append("")
            
            writer.writerow(row_data)
        
        # Write file2 data
        file2_rows = self.filter_rows(results_data["file2"]["rows"], include_matched, include_unchanged)
        for row in file2_rows:
            row_data = [
                "File2",
                row["row_number"],
                row["excel_row"],
                row["match_status"],
                row["has_changes"]
            ]
            
            # Add column values
            for header in results_data["file2"]["headers"]:
                if header in row["data"]:
                    row_data.append(row["data"][header]["value"])
                else:
                    row_data.append("")
            
            writer.writerow(row_data)
        
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="comparison_export.csv"'
        return response
    
    def export_as_excel(self, results_data, include_matched, include_unchanged, comparison):
        """Export as Excel format with separate sheets for each file"""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets for each file
        for file_key in ["file1", "file2"]:
            ws = wb.create_sheet(title=f"File{file_key[-1]}")
            file_data = results_data[file_key]
            
            # Write headers
            headers = ["Row_Number", "Excel_Row", "Match_Status", "Has_Changes"] + file_data["headers"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Write data
            filtered_rows = self.filter_rows(file_data["rows"], include_matched, include_unchanged)
            for row_idx, row in enumerate(filtered_rows, 2):
                ws.cell(row=row_idx, column=1, value=row["row_number"])
                ws.cell(row=row_idx, column=2, value=row["excel_row"])
                
                # Color-code match status
                status_cell = ws.cell(row=row_idx, column=3, value=row["match_status"])
                if row["match_status"] == "matched":
                    status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif row["match_status"] == "changed":
                    status_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif row["match_status"] == "added":
                    status_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                elif row["match_status"] == "removed":
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                ws.cell(row=row_idx, column=4, value=row["has_changes"])
                
                # Write column data
                for col_idx, header in enumerate(file_data["headers"], 5):
                    if header in row["data"]:
                        cell = ws.cell(row=row_idx, column=col_idx, value=row["data"][header]["value"])
                        
                        # Highlight changed columns
                        if row["data"][header].get("is_changed", False):
                            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        # Create summary sheet
        summary_ws = wb.create_sheet(title="Summary", index=0)
        
        # Add metadata
        summary_ws['A1'] = "Comparison Summary"
        summary_ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        summary_ws[f'A{row}'] = "Primary Column:"
        summary_ws[f'B{row}'] = results_data["comparison_metadata"].get("primary_column", "")
        
        row += 1
        summary_ws[f'A{row}'] = "Timestamp:"
        summary_ws[f'B{row}'] = results_data["comparison_metadata"].get("timestamp", "")
        
        row += 2
        summary_ws[f'A{row}'] = "File 1 Statistics:"
        summary_ws[f'A{row}'].font = Font(bold=True)
        
        file1_stats = results_data["file1"]["statistics"]
        for key, value in file1_stats.items():
            row += 1
            summary_ws[f'A{row}'] = key.replace("_", " ").title() + ":"
            summary_ws[f'B{row}'] = value
        
        row += 2
        summary_ws[f'A{row}'] = "File 2 Statistics:"
        summary_ws[f'A{row}'].font = Font(bold=True)
        
        file2_stats = results_data["file2"]["statistics"]
        for key, value in file2_stats.items():
            row += 1
            summary_ws[f'A{row}'] = key.replace("_", " ").title() + ":"
            summary_ws[f'B{row}'] = value
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="comparison_export.xlsx"'
        return response























